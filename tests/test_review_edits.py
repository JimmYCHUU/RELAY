"""Input notes that stay true, and the three hand edits the review screen offers.

The notes half of this file exists because a real July cycle put 29 per-cell
notes on screen about 27 cells, 26 of which the Facebook pass had since filled.
Each note was accurate when it was written and none was ever retracted, so the
panel read as a list of live problems that were mostly not problems.
"""
import openpyxl
import pytest
from fastapi.testclient import TestClient

from relay import store
from relay.models import SLOTS, CellValue, ReportRow, RowIssue, RunResult
from relay.resolve.insights_fill import row_issue
from relay.web.app import _live_issues


def make_result(n=3, brand="Brand C"):
    rows = []
    for i in range(n):
        links = {"fb1": f"https://facebook.com/somoynews.tv/posts/p{i}",
                 "fb2": f"https://facebook.com/somoytvsports/posts/p{i}",
                 "fb3": None, "x": None, "ig": None}
        cells = {s: CellValue.missing("awaiting the insights export") for s in SLOTS}
        rows.append(ReportRow(no=i + 1, date=None, caption=f"post {i}",
                              links=links, cells=cells))
    return RunResult(brand=brand, month="July", rows=rows)


# ── notes that stop being true ────────────────────────────────────────────────
def test_a_note_is_dropped_once_its_cell_has_a_value():
    run = make_result()
    row_issue(run, run.rows[0], ["fb1"], "no post on this page matches", row_idx=0)
    assert len(_live_issues(run)) == 1

    run.rows[0].cells["fb1"] = CellValue(66783, "collected", 1.0, "post id")
    assert _live_issues(run) == [], \
        "a filled cell retracts the note explaining why it was empty"


def test_a_note_is_dropped_once_its_link_is_gone():
    run = make_result()
    row_issue(run, run.rows[1], ["fb2"], "the post was reached but its id is "
                                         "in none of the supplied exports", row_idx=1)
    run.rows[1].links["fb2"] = None
    assert _live_issues(run) == []


def test_the_newest_reason_for_a_slot_replaces_the_older_one():
    """Three passes try each Facebook cell in turn and each has its own reason
    for giving up. Appending all three left one cell wearing three contradictory
    explanations, only the last of which still applied."""
    run = make_result()
    row_issue(run, run.rows[0], ["fb1"], "this page ran the same story twice", row_idx=0)
    row_issue(run, run.rows[0], ["fb1"], "two posts carry the same story", row_idx=0)
    row_issue(run, run.rows[0], ["fb1"], "the post was reached but its id is "
                                         "in none of the supplied exports", row_idx=0)
    live = _live_issues(run)
    assert len(live) == 1
    assert "in none of the supplied exports" in live[0].reason


def test_notes_for_different_slots_and_rows_coexist():
    run = make_result()
    row_issue(run, run.rows[0], ["fb1", "fb2"], "link names no page", row_idx=0)
    row_issue(run, run.rows[1], ["fb1"], "link names no page", row_idx=1)
    assert len(_live_issues(run)) == 3
    assert {i.slot for i in _live_issues(run)} == {"fb1", "fb2"}


def test_a_note_about_the_file_rather_than_a_cell_always_stands():
    run = make_result()
    run.issues.append(RowIssue("campaign.xlsx", 5, "empty Date filled from adjacent row",
                               where="sheet row 5"))
    run.issues.append(RowIssue("insights exports", 0, "12 links point at a page "
                                                     "none of the exports cover"))
    for row in run.rows:                       # every cell resolved
        for slot in ("fb1", "fb2"):
            row.cells[slot] = CellValue(10, "collected", 1.0, "export")
    assert len(_live_issues(run)) == 2


def test_a_row_number_says_which_kind_of_row_it_is():
    """"row 5" means a line of the workbook to the ingest step and the No the
    review table shows to the resolve step, and the panel listed both."""
    run = make_result()
    row_issue(run, run.rows[2], ["fb1"], "no post matches", row_idx=2)
    sheet_note = RowIssue("campaign.xlsx", 41, "empty Date filled", where="sheet row 41")
    assert run.issues[0].where == "row 3"       # the No the table shows
    assert sheet_note.where == "sheet row 41"
    assert RowIssue("insights exports", 0, "whole-file note").where == "", \
        "row 0 marks a note about no particular row"


# ── the web surface ───────────────────────────────────────────────────────────
def _campaign_sheet(tmp_path):
    """A two-row campaign tab in the layout the parser expects."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "July"
    ws["A1"] = "internal note"
    for col, head in enumerate(
            ["No", "Date", "Content's name", "Content's Link 1", "Content's Link 2",
             "Content's Link 3", "X, Link 4", "Instagram"], start=1):
        ws.cell(2, col, head)
    for i in (0, 1):
        ws.cell(3 + i, 1, i + 1)
        ws.cell(3 + i, 3, f"caption {i}")
        ws.cell(3 + i, 4, f"https://www.facebook.com/somoynews.tv/posts/p{i}")
        ws.cell(3 + i, 7, f"https://x.com/somoytv/status/{i}")
    path = tmp_path / "campaign.xlsx"
    wb.save(path)
    wb.close()
    return path


@pytest.fixture()
def client(tmp_path, monkeypatch):
    from relay import config
    monkeypatch.setattr(config, "DATA_DIR", tmp_path)
    monkeypatch.setattr(config, "OUTPUT_DIR", tmp_path / "output")
    monkeypatch.setattr(config, "DB_PATH", tmp_path / "db" / "runs.db")
    from relay.web import app as webapp
    monkeypatch.setattr(webapp, "UPLOADS", tmp_path / "uploads")
    monkeypatch.setattr(webapp, "_runs", {})
    monkeypatch.setattr(webapp, "_run_db_ids", {})
    monkeypatch.setattr(webapp, "_run_inputs", {})
    monkeypatch.setattr(webapp, "_jobs", {})
    monkeypatch.setattr(webapp, "_batch_jobs", {})
    monkeypatch.setattr(webapp, "_autopilot_job", None)
    return TestClient(webapp.app)


@pytest.fixture()
def run(client, tmp_path):
    res = client.post("/api/run", json={"campaign": str(_campaign_sheet(tmp_path)),
                                        "sheet": "July", "brand": "Brand C"})
    assert res.status_code == 200, res.text
    return res.json()


def test_manual_entry_carries_reach_and_engagement(client, run):
    """Meta publishes all three figures per Facebook post and the report has a
    column for each, so a hand-typed cell can carry all three too."""
    res = client.post("/api/override", json={
        "run_id": run["run_id"], "row_no": 1, "slot": "fb1",
        "value": 147461, "reach": 120300, "engagement": 4820})
    assert res.status_code == 200, res.text
    cell = res.json()
    assert (cell["value"], cell["reach"], cell["engagement"]) == (147461, 120300, 4820)
    assert cell["provenance"] == "manual"
    assert "reach 120,300" in cell["note"] and "engagement 4,820" in cell["note"]


def test_hand_typed_figures_reach_the_delivered_workbook(client, run, tmp_path):
    """The Reach and Engagement columns are the point of typing them in."""
    client.post("/api/override", json={
        "run_id": run["run_id"], "row_no": 1, "slot": "fb1",
        "value": 147461, "reach": 120300, "engagement": 4820})
    assert client.post(f"/api/report/{run['run_id']}").status_code == 200
    dl = client.get(f"/api/report/{run['run_id']}/download")
    out = tmp_path / "report.xlsx"
    out.write_bytes(dl.content)
    wb = openpyxl.load_workbook(out)
    ws = wb["July"]
    # row 3 is the first data row; E/F/G are FB 1's Views / Reach / Engagement
    assert (ws["E3"].value, ws["F3"].value, ws["G3"].value) == (147461, 120300, 4820)
    wb.close()


def test_manual_entry_survives_a_re_run(client, run, tmp_path):
    """The three figures are checkpointed together — a resume that brought back
    only the view count silently emptied the report's other two columns."""
    client.post("/api/override", json={
        "run_id": run["run_id"], "row_no": 2, "slot": "fb1",
        "value": 5000, "reach": 4000, "engagement": 300})
    again = client.post("/api/run", json={"campaign": str(tmp_path / "campaign.xlsx"),
                                          "sheet": "July", "brand": "Brand C"}).json()
    cell = again["rows"][1]["cells"]["fb1"]
    assert (cell["value"], cell["reach"], cell["engagement"]) == (5000, 4000, 300)


def test_reach_is_refused_where_the_report_has_no_column_for_it(client, run):
    res = client.post("/api/override", json={
        "run_id": run["run_id"], "row_no": 1, "slot": "x",
        "value": 900, "reach": 800})
    assert res.status_code == 400
    assert "no reach or engagement column" in res.json()["detail"]


def test_a_dead_link_can_be_struck_off_the_sheet(client, run, tmp_path):
    """A post gets taken down and nobody removes the link, so the cell sits
    there as work no collector can ever finish."""
    rid = run["run_id"]
    res = client.post("/api/link/remove",
                      json={"run_id": rid, "row_no": 1, "slot": "fb1"})
    assert res.status_code == 200, res.text
    assert res.json()["removed"].endswith("/p0")
    assert "no longer on the platform" in res.json()["cell"]["note"]

    # …and the workbook points the sponsor at nothing rather than at a dead page
    client.post(f"/api/report/{rid}")
    out = tmp_path / "report.xlsx"
    out.write_bytes(client.get(f"/api/report/{rid}/download").content)
    wb = openpyxl.load_workbook(out)
    ws = wb["July"]
    assert ws["D3"].value is None and ws["E3"].value is None
    assert ws["D4"].value.endswith("/p1"), "the row below keeps its own link"
    wb.close()


def test_striking_a_link_stops_it_counting_as_missing(client, run):
    from relay.web.app import _get_run
    rid = run["run_id"]
    assert _get_run(rid).coverage()["fb1"] == 0.0     # two links, neither filled
    client.post("/api/link/remove", json={"run_id": rid, "row_no": 1, "slot": "fb1"})
    client.post("/api/override", json={"run_id": rid, "row_no": 2, "slot": "fb1",
                                       "value": 10})
    assert _get_run(rid).coverage()["fb1"] == 1.0, \
        "a struck link drops out of the slot's denominator instead of failing it"


def test_a_struck_link_stays_struck_on_a_re_run(client, run, tmp_path):
    """The campaign sheet still lists the post — nobody went back and deleted
    the link, which is the whole reason the button exists."""
    client.post("/api/link/remove",
                json={"run_id": run["run_id"], "row_no": 1, "slot": "fb1"})
    again = client.post("/api/run", json={"campaign": str(tmp_path / "campaign.xlsx"),
                                          "sheet": "July", "brand": "Brand C"}).json()
    assert again["links_removed"] == 1
    assert again["rows"][0]["links"]["fb1"] is None
    assert "no longer on the platform" in again["rows"][0]["cells"]["fb1"]["note"]


def test_removing_a_link_that_is_not_there_is_refused(client, run):
    res = client.post("/api/link/remove",
                      json={"run_id": run["run_id"], "row_no": 1, "slot": "ig"})
    assert res.status_code == 409


def test_a_campaign_can_be_dropped_out_of_the_cycle(client, run):
    """The wrong sheet is only recognisable once its rows are on screen."""
    rid = run["run_id"]
    res = client.delete(f"/api/run/{rid}")
    assert res.status_code == 200
    assert res.json() == {"discarded": rid, "remaining": 0}
    assert client.get(f"/api/collect/{rid}/fb").status_code == 404
    assert client.delete(f"/api/run/{rid}").status_code == 404
    assert [r["status"] for r in client.get("/api/runs").json()] == ["discarded"]


def test_a_campaign_cannot_be_dropped_mid_collection(client, run, monkeypatch):
    """A running collector holds its own reference to the rows and would keep
    writing into a campaign the user believes is gone."""
    from relay.collectors.runner import Progress
    from relay.web import app as webapp
    monkeypatch.setitem(webapp._batch_jobs, "fb", Progress())   # state: running
    res = client.delete(f"/api/run/{run['run_id']}")
    assert res.status_code == 409
    assert "stop it first" in res.json()["detail"]


def test_an_override_lands_only_on_the_row_it_names(tmp_path):
    """A campaign sheet's No is hand-filled and may repeat, so a write keyed on
    the number alone landed in every row that shared it."""
    db = tmp_path / "runs.db"
    result = make_result(2)
    for row in result.rows:
        row.no = 7                              # the sheet repeats a No
    run_id = store.save_run(result, {"campaign": "c.xlsx", "sheet": "July",
                                     "brand": "Brand C"}, db_path=db)
    cell = CellValue(99, "manual", 1.0, "typed in by hand", reach=80, engagement=5)
    store.record_override(run_id, 7, "fb1", None, 99, db_path=db, cell=cell, row_idx=1)
    got = {c["row_idx"]: c["value"] for c in store.load_cells(run_id, db_path=db)}
    assert got == {1: 99}
