"""Per-brand report colour, taken from the campaign sheet (FR-20 visual layer).

Fixtures are built in-process: the real sponsor trackers live on the analyst's
machine, not in the repo.
"""
import openpyxl
import pytest
from openpyxl.styles import PatternFill

from relay.report.generator import ACCENT, BAND, TINT, build_report
from relay.report.palette import (DEFAULT, Palette, accent_from_campaign,
                                  contrast, derive, is_brand_colour,
                                  mix_on_white, readable_on)
from relay.models import CellValue, ReportRow, RunResult

# The four colours the real sponsor trackers actually use.
BKASH, TK, FRESH, SMC = "A64D79", "F1C232", "1155CC", "93C47D"


def _campaign(tmp_path, fill_hex: str | None, name="camp.xlsx"):
    """A campaign sheet shaped like the real ones: brand name on row 1, header
    on row 2, both filled with the brand's colour."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "June"
    ws["A1"] = "BRAND"
    headers = ["No", "Date", "Content's name", "Content's Link 1",
               "Content's Link 2", "Content's Link 3", "X, Link 4", "Instagram"]
    for i, h in enumerate(headers, start=1):
        ws.cell(row=2, column=i, value=h)
    if fill_hex:
        fill = PatternFill("solid", fgColor="FF" + fill_hex)
        for col in range(1, 9):
            ws.cell(row=1, column=col).fill = fill
            ws.cell(row=2, column=col).fill = fill
    ws["A3"], ws["C3"] = 1, "a caption"
    ws["D3"] = "https://www.facebook.com/x/posts/1"
    path = tmp_path / name
    wb.save(path)
    wb.close()
    return path


# ── derivation ────────────────────────────────────────────────────────────────
@pytest.mark.parametrize("accent,tint,band", [
    ("1F3864", "E9EBF0", "F6F7F9"),   # style-6 "fallback navy"
    ("C0392B", "F9EBEA", "FCF7F7"),   # style-6 "demo brand color"
])
def test_derive_reproduces_approved_samples(accent, tint, band):
    """The 10%/4% mixes are not invented — they reproduce both style-6 samples
    exactly, which is what makes them the house formula."""
    pal = derive(accent)
    assert pal.tint == tint
    assert pal.band == band


def test_default_is_byte_identical_to_the_approved_teal():
    """An unbranded sheet must produce exactly today's approved workbook."""
    assert DEFAULT.accent == ACCENT
    assert DEFAULT.tint == TINT
    assert DEFAULT.band == BAND
    assert DEFAULT.on_accent == "FFFFFF"
    assert DEFAULT.accent_text == ACCENT
    assert DEFAULT.rule == ACCENT


def test_mix_on_white_endpoints():
    assert mix_on_white("123456", 0.0) == "FFFFFF"
    assert mix_on_white("123456", 1.0) == "123456"


# ── legibility ────────────────────────────────────────────────────────────────
@pytest.mark.parametrize("accent,expect", [
    (TK, "1F2937"),      # gold: white 38pt on it is unreadable
    (SMC, "1F2937"),     # light green: same
    (BKASH, "FFFFFF"),
    (FRESH, "FFFFFF"),
])
def test_banner_ink_flips_on_light_accents(accent, expect):
    assert readable_on(accent) == expect


@pytest.mark.parametrize("accent", [BKASH, TK, FRESH, SMC])
def test_every_real_brand_clears_large_text_contrast(accent):
    """38pt bold banner is WCAG 'large text': AA is 3:1."""
    pal = derive(accent)
    assert contrast(pal.on_accent, pal.accent) >= 3.0


@pytest.mark.parametrize("accent", [BKASH, TK, FRESH, SMC])
def test_sum_figures_stay_readable_on_their_own_tint(accent):
    """Gold Sum figures on a pale gold tint would vanish undarkened."""
    pal = derive(accent)
    assert contrast(pal.accent_text, pal.tint) >= 4.5


# ── what counts as a brand colour ─────────────────────────────────────────────
@pytest.mark.parametrize("hexv", [BKASH, TK, FRESH, SMC])
def test_real_brand_colours_are_accepted(hexv):
    assert is_brand_colour(hexv)


@pytest.mark.parametrize("hexv", [
    "CCCCCC",   # the unbranded tracker's grey
    "FFFFFF", "000000", "F2F2F2", "808080",
])
def test_neutrals_are_not_brand_colours(hexv):
    assert not is_brand_colour(hexv)


# ── extraction ────────────────────────────────────────────────────────────────
@pytest.mark.parametrize("hexv", [BKASH, TK, FRESH, SMC])
def test_accent_read_from_header_rows(tmp_path, hexv):
    assert accent_from_campaign(_campaign(tmp_path, hexv), "June") == hexv


def test_unbranded_sheet_yields_no_accent(tmp_path):
    assert accent_from_campaign(_campaign(tmp_path, "CCCCCC"), "June") is None
    assert accent_from_campaign(_campaign(tmp_path, None, "b.xlsx"), "June") is None


def test_highlighted_data_cell_never_becomes_the_masthead(tmp_path):
    """The olive-masthead regression: a yellow highlight far below the header
    must not be mistaken for the brand."""
    path = _campaign(tmp_path, None)
    wb = openpyxl.load_workbook(path)
    ws = wb["June"]
    for r in range(3, 40):
        ws.cell(row=r, column=3).fill = PatternFill("solid", fgColor="FFFFFF00")
    wb.save(path)
    wb.close()
    assert accent_from_campaign(path, "June") is None


def test_missing_sheet_and_bad_file_fall_back(tmp_path):
    assert accent_from_campaign(_campaign(tmp_path, BKASH), "Nope") is None
    bad = tmp_path / "not.xlsx"
    bad.write_text("plainly not a workbook")
    assert accent_from_campaign(bad, "June") is None


# ── the generated workbook ────────────────────────────────────────────────────
def _run(accent):
    row = ReportRow(no=1, date=None, caption="c",
                    links={s: None for s in ("fb1", "fb2", "fb3", "x", "ig")},
                    cells={s: CellValue(1, "manual") for s in ("fb1", "fb2", "fb3", "x", "ig")})
    return RunResult(brand="B", month="June", rows=[row], accent=accent)


def test_chrome_wears_the_brand_but_data_rows_do_not(tmp_path):
    out = build_report(_run(BKASH), tmp_path / "r.xlsx")
    ws = openpyxl.load_workbook(out).active
    pal = derive(BKASH)
    assert ws["A1"].fill.fgColor.rgb.endswith(pal.accent)          # banner
    assert ws["A2"].fill.fgColor.rgb.endswith(pal.tint)            # header row
    assert ws["A2"].border.bottom.color.rgb.endswith(pal.rule)     # the divider
    sum_row = 3 + 1
    assert ws[f"A{sum_row}"].fill.fgColor.rgb.endswith(pal.tint)   # Sum
    assert ws[f"A{sum_row + 1}"].fill.fgColor.rgb.endswith(pal.accent)  # Total
    assert ws[f"A{sum_row + 2}"].fill.fgColor.rgb.endswith(pal.accent)  # Average
    # the rows a reader actually reads stay neutral
    assert ws["A3"].fill.fgColor.rgb.endswith(BAND) or ws["A3"].fill.fill_type is None


def test_unbranded_run_still_produces_the_teal_report(tmp_path):
    out = build_report(_run(None), tmp_path / "r.xlsx")
    ws = openpyxl.load_workbook(out).active
    assert ws["A1"].fill.fgColor.rgb.endswith(ACCENT)
    assert ws["A2"].fill.fgColor.rgb.endswith(TINT)


def test_each_brand_in_a_cycle_keeps_its_own_colour(tmp_path):
    """A multi-brand run writes one workbook per brand; they must not share a
    palette."""
    seen = {}
    for name, accent in (("bkash", BKASH), ("tk", TK), ("fresh", FRESH)):
        out = build_report(_run(accent), tmp_path / f"{name}.xlsx")
        ws = openpyxl.load_workbook(out).active
        seen[name] = ws["A1"].fill.fgColor.rgb[-6:]
    assert seen == {"bkash": BKASH, "tk": TK, "fresh": FRESH}
