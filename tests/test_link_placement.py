"""Links that are in the wrong column, and the gap that leaves at the head of a row.

Two separate rules, at two different stages:

  * ingest puts a link in the column its own URL names, because a crossed link
    is not a wrong figure but a permanently empty cell;
  * the report closes up the first Facebook column, because a sponsor reads an
    empty Link 1 beside a filled Link 2 as a hole in the report.
"""
import openpyxl
import pytest

from relay.ingest.campaign import parse_campaign
from relay.models import SLOTS, CellValue, ReportRow, RunResult
from relay.report.generator import _fb_order, build_report

FB = "https://www.facebook.com/somoynews.tv/posts/pfbid0{}AbCdEf"
FB2 = "https://www.facebook.com/somoytvsports/posts/pfbid0{}GhIjKl"
X = "https://x.com/somoytv/status/207155236188212432{}"
IG = "https://www.instagram.com/p/DM{}xyzAB/"
PROFILE = "https://www.facebook.com/somoynews.tv/"
SHARE = "https://www.facebook.com/share/p/1AbCdEfGhi/"

HEADER = ["No", "Date", "Content's name", "Content's Link [1]", "Content's Link 2",
          "Content's Link 3", "X, Link 4", "Instagram"]


def _sheet(tmp_path, rows, name="c.xlsx"):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "July"
    for c, v in enumerate(HEADER, start=1):
        ws.cell(1, c, v)
    for r, values in enumerate(rows, start=2):
        for c, v in enumerate(values, start=1):
            ws.cell(r, c, v)
    path = tmp_path / name
    wb.save(path)
    wb.close()
    return path


# ── ingest: a link belongs to the column its URL names ────────────────────────
def test_a_crossed_x_and_instagram_pair_is_swapped_back(tmp_path):
    """The two columns sit side by side in a hand-filled sheet."""
    path = _sheet(tmp_path, [[1, None, "cap", FB.format(1), None, None,
                              IG.format("a"), X.format(1)]])
    rows, issues = parse_campaign(path, "July")
    assert rows[0].x_link == X.format(1)
    assert rows[0].ig_link == IG.format("a")
    moves = [i.reason for i in issues if "read as" in i.reason]
    assert moves == [
        "the X column held an Instagram link — read as the Instagram column",
        "the Instagram column held an X link — read as the X column",
    ]


def test_a_stray_instagram_link_moves_into_an_empty_instagram_column(tmp_path):
    """bKash's July row 120: an Instagram URL pasted into a Facebook column."""
    path = _sheet(tmp_path, [[1, None, "cap", FB.format(1), IG.format("a"), None,
                              X.format(1), None]])
    rows, issues = parse_campaign(path, "July")
    assert rows[0].fb_links == [FB.format(1), None, None]
    assert rows[0].ig_link == IG.format("a")
    assert any("Link 2 held an Instagram link — read as the Instagram column"
               == i.reason for i in issues)


def test_a_stray_link_with_nowhere_to_go_stays_where_the_sheet_put_it(tmp_path):
    """Moving it over the Instagram post already there would lose a link the
    sponsor supplied, so it is reported rather than resolved."""
    path = _sheet(tmp_path, [[1, None, "cap", FB.format(1), IG.format("a"), None,
                              None, IG.format("b")]])
    rows, issues = parse_campaign(path, "July")
    assert rows[0].fb_links[1] == IG.format("a"), "not dropped"
    assert rows[0].ig_link == IG.format("b"), "and not overwritten"
    assert not any("read as" in i.reason for i in issues)
    assert any("already taken by another post" in i.reason for i in issues)


def test_the_same_post_pasted_into_two_columns_is_only_delivered_once(tmp_path):
    """bKash's July row 120: one Instagram reel in both the Instagram column and
    a Facebook link column, where it can only ever print with blank figures."""
    path = _sheet(tmp_path, [[1, None, "cap", FB.format(1), IG.format("a"), None,
                              None, IG.format("a")]])
    rows, issues = parse_campaign(path, "July")
    assert rows[0].fb_links == [FB.format(1), None, None], "the stray copy goes"
    assert rows[0].ig_link == IG.format("a"), "the one in its own column stays"
    assert any("this copy is left out" in i.reason for i in issues)


def test_a_double_paste_is_recognised_through_instagram_s_tracking_suffix(tmp_path):
    tracked = IG.format("a") + "?utm_source=ig_web_copy_link&igsh=MzRlODBiNWFlZA=="
    path = _sheet(tmp_path, [[1, None, "cap", FB.format(1), tracked, None,
                              None, IG.format("a")]])
    rows, _ = parse_campaign(path, "July")
    assert rows[0].fb_links[1] is None
    assert rows[0].ig_link == IG.format("a")


def test_links_already_in_the_right_columns_are_left_alone(tmp_path):
    path = _sheet(tmp_path, [[1, None, "cap", FB.format(1), FB2.format(2),
                              None, X.format(1), IG.format("a")]])
    rows, issues = parse_campaign(path, "July")
    assert rows[0].fb_links == [FB.format(1), FB2.format(2), None]
    assert (rows[0].x_link, rows[0].ig_link) == (X.format(1), IG.format("a"))
    assert not any("read as" in i.reason for i in issues)


def test_a_facebook_link_in_the_x_column_comes_back_to_a_free_link_column(tmp_path):
    path = _sheet(tmp_path, [[1, None, "cap", FB.format(1), None, None,
                              FB2.format(2), None]])
    rows, _ = parse_campaign(path, "July")
    assert rows[0].fb_links == [FB.format(1), FB2.format(2), None]
    assert rows[0].x_link is None


# ── report: the first Facebook column is never the empty one ──────────────────
def _row(links):
    slots = dict(zip(("fb1", "fb2", "fb3"), links))
    return ReportRow(no=1, date=None, caption="cap",
                     links={**slots, "x": None, "ig": None},
                     cells={s: CellValue.missing() for s in SLOTS})


def test_an_empty_first_link_is_filled_from_the_next_one():
    assert _fb_order(_row([None, FB.format(2), FB2.format(3)])) \
        == ["fb2", "fb1", "fb3"], "link 2 is promoted and vacates its own column"


def test_a_first_link_that_names_no_post_is_demoted():
    assert _fb_order(_row([PROFILE, FB.format(2), None])) == ["fb2", "fb1", "fb3"]


def test_a_share_link_counts_as_naming_a_post():
    assert _fb_order(_row([SHARE, FB.format(2), None])) == ["fb1", "fb2", "fb3"]


def test_a_gap_after_the_first_column_is_left_alone():
    assert _fb_order(_row([FB.format(1), None, FB2.format(3)])) \
        == ["fb1", "fb2", "fb3"], "only the head of the row is closed up"


def test_a_row_with_no_usable_link_at_all_keeps_its_order():
    assert _fb_order(_row([None, None, None])) == ["fb1", "fb2", "fb3"]
    assert _fb_order(_row([PROFILE, None, None])) == ["fb1", "fb2", "fb3"]


def _built(tmp_path, links, cells):
    rows = [ReportRow(no=1, date=None, caption="cap",
                      links={**dict(zip(("fb1", "fb2", "fb3"), links)),
                             "x": X.format(1), "ig": IG.format("a")},
                      cells=cells)]
    out = tmp_path / "r.xlsx"
    build_report(RunResult(brand="Brand", month="July", rows=rows), out)
    wb = openpyxl.load_workbook(out)
    ws = wb["July"]
    got = {c: ws.cell(3, i).value for c, i in
           (("L1", 4), ("V1", 5), ("R1", 6), ("E1", 7),
            ("L2", 8), ("V2", 9), ("L3", 12), ("V3", 13),
            ("X", 16), ("XV", 17), ("IG", 18), ("IGV", 19))}
    wb.close()
    return got


def test_the_promoted_link_takes_its_three_figures_with_it(tmp_path):
    cells = {s: CellValue.missing() for s in SLOTS}
    cells["fb2"] = CellValue(70512, "collected", 1.0, "export", reach=60000, engagement=1200)
    cells["x"] = CellValue(900, "collected", 1.0, "x")
    got = _built(tmp_path, [None, FB.format(2), None], cells)
    assert (got["L1"], got["V1"], got["R1"], got["E1"]) == (FB.format(2), 70512, 60000, 1200)
    assert (got["L2"], got["V2"]) == (None, None), "and vacates the column it came from"
    assert (got["X"], got["XV"]) == (X.format(1), 900), "X never moves"
    assert got["IG"] == IG.format("a"), "nor Instagram"


def test_a_full_row_is_written_exactly_as_the_sheet_has_it(tmp_path):
    cells = {s: CellValue.missing() for s in SLOTS}
    for slot, v in (("fb1", 10), ("fb2", 20), ("fb3", 30)):
        cells[slot] = CellValue(v, "collected", 1.0, "export")
    got = _built(tmp_path, [FB.format(1), FB2.format(2), FB.format(3)], cells)
    assert (got["L1"], got["V1"]) == (FB.format(1), 10)
    assert (got["L2"], got["V2"]) == (FB2.format(2), 20)
    assert (got["L3"], got["V3"]) == (FB.format(3), 30)


@pytest.mark.parametrize("links,expect_first", [
    ([None, FB.format(2), FB2.format(3)], FB.format(2)),
    ([None, None, FB2.format(3)], FB2.format(3)),
    ([PROFILE, FB.format(2), None], FB.format(2)),
])
def test_the_first_column_always_leads_with_a_real_post(tmp_path, links, expect_first):
    got = _built(tmp_path, links, {s: CellValue.missing() for s in SLOTS})
    assert got["L1"] == expect_first
    written = {got["L1"], got["L2"], got["L3"]} - {None}
    assert written == {l for l in links if l}, "every link the sheet had is still delivered"
