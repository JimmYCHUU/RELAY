import openpyxl
import pytest

from relay.report.generator import ACCENT, BAND, TINT, build_report


@pytest.fixture()
def generated(april_result, tmp_path):
    out = tmp_path / "out.xlsx"
    build_report(april_result, out)
    wb = openpyxl.load_workbook(out)
    yield wb["April"]
    wb.close()


def test_banner_and_header(generated):
    ws = generated
    assert "A1:V1" in {str(r) for r in ws.merged_cells.ranges}
    assert ws["A1"].value.startswith("BRAND A")
    assert ws["A1"].font.size == 38 and ws["A1"].font.bold
    # no campaign sheet given -> the theme's fallback accent
    assert ws["A1"].fill.fgColor.rgb.endswith(ACCENT)
    assert ws["A2"].value == "No" and ws["E2"].value == "Views"
    assert ws["A2"].font.size == 12 and ws["A2"].font.bold
    # each Facebook link carries Views / Reach / Engagement / Clicks; X and
    # Instagram, which no export reports the other three for, keep one column
    assert [ws.cell(row=2, column=c).value for c in range(4, 9)] == \
        ["Content's Link 1", "Views", "Reach", "Engagement", "Clicks"]
    assert [ws.cell(row=2, column=c).value for c in range(9, 14)] == \
        ["Content's Link 2", "Views", "Reach", "Engagement", "Clicks"]
    assert [ws.cell(row=2, column=c).value for c in (19, 20, 21, 22)] == \
        ["X, Link 4", "Impressions", "Instagram", "Views"]
    # a single-tab report says nothing about source tabs
    assert ws.cell(row=2, column=23).value is None


def test_footer_formulas(generated):
    ws = generated
    n = 25
    sum_row = n + 3
    views_row, reach_row, eng_row, clicks_row = (sum_row + i for i in (1, 2, 3, 4))
    assert ws.cell(row=sum_row, column=1).value == "Sum"
    assert ws.cell(row=sum_row, column=5).value == f"=SUM(E3:E{n + 2})"
    assert ws.cell(row=sum_row, column=22).value == f"=SUM(V3:V{n + 2})"
    # link columns carry no total
    assert ws.cell(row=sum_row, column=4).value is None
    # the four metrics interleave, so each total adds its own columns
    assert ws.cell(row=views_row, column=1).value == "Total views"
    assert ws.cell(row=views_row, column=4).value == \
        f"=E{sum_row}+J{sum_row}+O{sum_row}+T{sum_row}+V{sum_row}"
    assert ws.cell(row=reach_row, column=1).value == "Total reach (Facebook)"
    assert ws.cell(row=reach_row, column=4).value == \
        f"=F{sum_row}+K{sum_row}+P{sum_row}"
    assert ws.cell(row=eng_row, column=1).value == "Total engagement (Facebook)"
    assert ws.cell(row=eng_row, column=4).value == \
        f"=G{sum_row}+L{sum_row}+Q{sum_row}"
    assert ws.cell(row=clicks_row, column=1).value == "Total clicks (Facebook)"
    assert ws.cell(row=clicks_row, column=4).value == \
        f"=H{sum_row}+M{sum_row}+R{sum_row}"


def test_footer_averages_divide_their_own_total_by_the_row_count(generated):
    """Four totals, then the same four as an average per content — each average
    pointing at the total four rows above it, not at whatever row it lands on."""
    ws = generated
    n = 25
    sum_row = n + 3
    labels = ["Average views per content", "Average reach per content (Facebook)",
              "Average engagement per content (Facebook)",
              "Average clicks per content (Facebook)"]
    for i, label in enumerate(labels):
        avg_row = sum_row + 5 + i
        assert ws.cell(row=avg_row, column=1).value == label
        assert ws.cell(row=avg_row, column=4).value == f"=D{sum_row + 1 + i}/{n}"
        assert ws.cell(row=avg_row, column=4).number_format == "0"


def test_footer_merges(generated):
    ws = generated
    merges = {str(r) for r in ws.merged_cells.ranges}
    sum_row = 25 + 3
    assert f"A{sum_row}:C{sum_row}" in merges
    # four totals + four averages
    for rr in range(sum_row + 1, sum_row + 9):
        assert f"A{rr}:C{rr}" in merges and f"D{rr}:V{rr}" in merges


def test_data_row_styles(generated):
    ws = generated
    assert ws["B3"].number_format == "d\\ mmm"
    assert ws["C3"].font.name == "Arial" and ws["C3"].font.size == 11
    link_cell = ws["D3"]
    assert link_cell.font.color.rgb.endswith("1155CC")
    assert link_cell.hyperlink is not None
    assert ws["A3"].border.bottom.style == "thin"
    assert ws["E3"].number_format == "#,##0"


def test_modern_theme_polish(generated):
    ws = generated
    assert ws.freeze_panes == "A3"
    assert ws.sheet_view.showGridLines is False
    # alternating banding: row 4 is banded, row 3 is not
    assert ws["A4"].fill.fill_type == "solid"
    assert ws["A3"].fill.fill_type != "solid"
    # nothing beyond column V reaches the sponsor on a single-tab report
    assert ws.max_column <= 22
    # rows are tall enough for wrapped captions
    assert ws.row_dimensions[3].height >= 70


def test_column_widths(generated):
    ws = generated
    expected = {"A": 16.7, "C": 31.3, "V": 19.6}
    for col, w in expected.items():
        assert abs(ws.column_dimensions[col].width - w) < 0.1
    # the unused Source tab column is never sized on a single-tab report
    assert "W" not in ws.column_dimensions


def _colored_campaign(path, fill_hex):
    wb = openpyxl.Workbook()
    ws = wb.active
    from openpyxl.styles import PatternFill
    for col in range(1, 8):
        ws.cell(row=1, column=col, value="x").fill = \
            PatternFill("solid", fgColor=fill_hex)
    ws.cell(row=2, column=1, value="y").fill = PatternFill("solid", fgColor="FFCCCCCC")
    wb.save(path)
    return path



