"""Campaign tabs that do not match the one layout the parser used to assume.

Every case here is a real tab from the July 2026 cycle. Positions and column
names vary per sponsor and per month, and reading by position silently produced
a wrong report rather than an error — which is the worst way for a parser to be
wrong.
"""
import openpyxl
import pytest

from relay.ingest.campaign import parse_campaign

FB = "https://www.facebook.com/somoynews.tv/posts/pfbid0{}"
X = "https://x.com/somoytv/status/207155236188212432{}"
IG = "https://www.instagram.com/p/DM{}xyz/"


def _write(tmp_path, header, rows, name="c.xlsx", tab="July", note_row=None):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = tab
    if note_row:
        for c, v in enumerate(note_row, start=1):
            ws.cell(1, c, v)
    top = 2 if note_row else 1
    for c, v in enumerate(header, start=1):
        ws.cell(top, c, v)
    for r, values in enumerate(rows, start=top + 1):
        for c, v in enumerate(values, start=1):
            ws.cell(r, c, v)
    path = tmp_path / name
    wb.save(path)
    wb.close()
    return path


STANDARD = ["No", "Date", "Content's name", "Content's Link [1]", "Content's Link 2",
            "Content's Link 3", "X, Link 4", "Instagram"]


def test_the_standard_layout_is_unchanged(tmp_path):
    path = _write(tmp_path, STANDARD,
                  [[1, None, "cap", FB.format("a"), FB.format("b"), FB.format("c"),
                    X.format(1), IG.format("a")]])
    rows, _ = parse_campaign(path, "July")
    assert rows[0].fb_links == [FB.format("a"), FB.format("b"), FB.format("c")]
    assert rows[0].x_link == X.format(1)
    assert rows[0].ig_link == IG.format("a")


def test_a_first_column_named_sl_is_still_a_header(tmp_path):
    """Grameenphone's own workbook says "No" on its April tab and " SL" on May
    through August. Requiring the literal "No" left five months unreadable."""
    header = [" SL"] + STANDARD[1:]
    path = _write(tmp_path, header, [[1, None, "cap", FB.format("a"), None, None,
                                      None, None]])
    rows, _ = parse_campaign(path, "July")
    assert len(rows) == 1 and rows[0].fb_links[0] == FB.format("a")


def test_a_blank_first_column_is_still_a_header(tmp_path):
    header = [" "] + STANDARD[1:]
    path = _write(tmp_path, header, [[1, None, "cap", FB.format("a"), None, None,
                                      None, None]])
    rows, _ = parse_campaign(path, "July")
    assert len(rows) == 1


def test_a_blank_column_between_links_does_not_shift_every_column_after_it(tmp_path):
    """Tecno POVA's June tab. Reading by position took its Link 3 for the X post
    and its X post for the Instagram one, so a Facebook URL was delivered in the
    report's X column and 30 Instagram cells came back empty against an export
    that held every one of them."""
    header = ["No", "Date", "Content's name", "Content's Link", None,
              "Content's Link 2", "Content's Link 3", "Content's Link 4 X",
              "Instagram", "Total Count"]
    path = _write(tmp_path, header,
                  [[1, None, "cap", FB.format("a"), None, FB.format("b"),
                    FB.format("c"), X.format(1), IG.format("a"), "Bioscope"]])
    rows, _ = parse_campaign(path, "July")
    r = rows[0]
    assert r.fb_links == [FB.format("a"), FB.format("b"), FB.format("c")]
    assert r.x_link == X.format(1), "the X column is the one the header names X"
    assert r.ig_link == IG.format("a"), "the Instagram column is read at all"


def test_a_tab_with_no_instagram_column_gets_no_instagram_link(tmp_path):
    """Grameenphone's tabs end at "Total Count" — reading the column after X on
    faith only worked because that cell happened not to look like a URL."""
    header = ["No", "Date", "Content's name", "Content's Link", "Content's Link 2",
              "Content's Link 3", "Content's Link 4 X", "Total Count"]
    path = _write(tmp_path, header,
                  [[1, None, "cap", FB.format("a"), FB.format("b"), FB.format("c"),
                    X.format(1), "Online Recharge"]])
    rows, _ = parse_campaign(path, "July")
    assert rows[0].x_link == X.format(1)
    assert rows[0].ig_link is None


def test_a_count_row_that_names_a_category_in_its_No_cell_is_dropped(tmp_path):
    """Every Grameenphone tab opens with one ("GP Bundle | | 46 | Online
    Recharge | | 6"), and it used to survive as a content row captioned "46"."""
    path = _write(tmp_path, STANDARD, [
        ["GP Bundle", None, 46, "Online Recharge", None, 6, "<OFF", None],
        [1, None, "a real caption", FB.format("a"), None, None, None, None],
    ])
    rows, issues = parse_campaign(path, "July")
    assert [r.caption for r in rows] == ["a real caption"]
    assert any("category-count row ignored" in i.reason for i in issues)


def test_a_numbered_row_with_no_links_is_never_mistaken_for_a_count_row(tmp_path):
    path = _write(tmp_path, STANDARD,
                  [[1, None, "a caption someone has yet to paste links for",
                    "pending", None, None, None, None]])
    rows, _ = parse_campaign(path, "July")
    assert len(rows) == 1, "a numbered row is content, however empty its links"


def test_a_header_the_mapper_cannot_read_falls_back_to_fixed_positions(tmp_path):
    """Two link columns is the floor; below it the old positions are safer than
    a half-built mapping."""
    header = ["No", "Date", "Content's name", "Content's Link", "Views", "Reach",
              "X, Link 4", "Instagram"]
    path = _write(tmp_path, header, [[1, None, "cap", FB.format("a"), None, None,
                                      X.format(1), IG.format("a")]])
    rows, _ = parse_campaign(path, "July")
    assert rows[0].fb_links[0] == FB.format("a")
    assert rows[0].x_link == X.format(1)


@pytest.mark.parametrize("first", ["No", "no.", "SL", "Sl.", "S. No", "Serial", "#", ""])
def test_row_number_column_spellings(tmp_path, first):
    path = _write(tmp_path, [first] + STANDARD[1:],
                  [[1, None, "cap", FB.format("a"), None, None, None, None]])
    assert len(parse_campaign(path, "July")[0]) == 1
