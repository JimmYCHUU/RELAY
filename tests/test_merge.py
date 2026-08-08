"""Several tabs of one brand's workbook, delivered as one report.

Ruchi's June campaign runs as three tabs — `June`, `June ratio 2` and
`8 Teams Special` — and the sponsor is handed one file. Cocola's tabs merge into
a second file, never into Ruchi's.
"""
from __future__ import annotations

from datetime import datetime

import openpyxl
import pytest
from fastapi.testclient import TestClient

from relay.models import CellValue, ReportRow, RunResult
from relay.report.generator import build_report
from relay.report.merge import merge_runs

FB = "https://www.facebook.com/somoynews.tv/posts/pfbid0"


def _run(brand: str, month: str, n: int, start: int = 1, accent=None) -> RunResult:
    rows = []
    for i in range(start, start + n):
        rows.append(ReportRow(
            no=i, date=datetime(2026, 6, 1 + (i % 20)), caption=f"{month} row {i}",
            links={"fb1": f"{FB}{i}", "fb2": None, "fb3": None, "x": None, "ig": None},
            cells={"fb1": CellValue(100 * i, "collected", 1.0, "export",
                                    reach=60 * i, engagement=3 * i,
                                    clicks=12 * i),
                   **{s: CellValue.missing() for s in ("fb2", "fb3", "x", "ig")}},
        ))
    return RunResult(brand=brand, month=month, rows=rows, accent=accent)


def test_tabs_merge_into_one_run_in_order():
    runs = [_run("Ruchi", "June", 3), _run("Ruchi", "June ratio 2", 2),
            _run("Ruchi", "8 Teams Special", 1)]
    merged = merge_runs(runs, "June")

    assert merged.brand == "Ruchi" and merged.month == "June"
    assert len(merged.rows) == 6
    assert [r.source_sheet for r in merged.rows] == \
        ["June"] * 3 + ["June ratio 2"] * 2 + ["8 Teams Special"]


def test_merging_copies_rows_and_leaves_the_source_runs_alone():
    """The review screen, the collectors and the resume checkpoints all key on
    the per-tab runs, so regenerating a merged report must not touch them."""
    runs = [_run("Ruchi", "June", 2), _run("Ruchi", "June ratio 2", 2)]
    merged = merge_runs(runs)

    assert all(r.source_sheet == "" for run in runs for r in run.rows)
    assert merged.rows[0] is not runs[0].rows[0]


def test_a_lone_tab_is_returned_unchanged():
    run = _run("Ruchi", "June", 2)
    assert merge_runs([run]) is run
    # …and named, it still carries no Source column: there is nothing to say
    assert all(r.source_sheet == "" for r in merge_runs([run], "June 2026").rows)


def test_the_first_branded_tab_speaks_for_the_file():
    runs = [_run("Ruchi", "June", 1), _run("Ruchi", "June ratio 2", 1, accent="93C47D")]
    assert merge_runs(runs, "June").accent == "93C47D"
    runs[0].accent = "FF0000"
    assert merge_runs(runs, "June").accent == "FF0000"


def test_merging_nothing_is_an_error():
    with pytest.raises(ValueError):
        merge_runs([])


def test_the_merged_workbook_renumbers_end_to_end_and_names_each_tab(tmp_path):
    runs = [_run("Ruchi", "June", 3), _run("Ruchi", "June ratio 2", 2)]
    out = build_report(merge_runs(runs, "June"), tmp_path / "Ruchi (June).xlsx")

    wb = openpyxl.load_workbook(out)
    ws = wb["June"]
    assert ws.max_column == 23                      # 22 + the Source tab column
    assert ws.cell(row=2, column=23).value == "Source tab"
    assert [ws.cell(row=3 + i, column=1).value for i in range(5)] == [1, 2, 3, 4, 5]
    assert [ws.cell(row=3 + i, column=23).value for i in range(5)] == \
        ["June"] * 3 + ["June ratio 2"] * 2
    # one set of totals, over all five rows
    sum_row = 3 + 5
    assert ws.cell(row=sum_row, column=1).value == "Sum"
    assert ws.cell(row=sum_row, column=5).value == "=SUM(E3:E7)"
    # the first average sits below all four totals and divides the first of them
    assert ws.cell(row=sum_row + 5, column=4).value == f"=D{sum_row + 1}/5"
    # the banner spans the wider sheet
    assert "A1:W1" in {str(r) for r in ws.merged_cells.ranges}
    wb.close()


def test_facebook_figures_land_in_their_own_columns(tmp_path):
    out = build_report(_run("Ruchi", "June", 1), tmp_path / "one.xlsx")
    wb = openpyxl.load_workbook(out)
    ws = wb["June"]
    # row 1 of the fixture: views 100, reach 60, engagement 3, clicks 12
    assert [ws.cell(row=3, column=c).value for c in (5, 6, 7, 8)] == [100, 60, 3, 12]
    # an unlinked slot leaves all of its columns empty
    assert [ws.cell(row=3, column=c).value for c in (10, 11, 12, 13)] == [None] * 4
    wb.close()


# --- the API surface the dashboard drives ---

def _staged(client, runs):
    """Push RunResults into the app's in-memory run table the way /api/run
    would, and return their ids."""
    from relay.web import app as webapp

    ids = []
    for i, run in enumerate(runs):
        rid = f"test{i}"
        webapp._runs[rid] = run
        webapp._run_db_ids[rid] = -1
        ids.append(rid)
    return ids


@pytest.fixture()
def client(tmp_path, monkeypatch):
    from relay import config, store
    from relay.web import app as webapp

    monkeypatch.setattr(config, "OUTPUT_DIR", tmp_path)
    monkeypatch.setattr(store, "set_output", lambda *a, **k: None)
    webapp._runs.clear()
    webapp._run_db_ids.clear()
    return TestClient(webapp.app)


def test_report_group_endpoint_writes_one_workbook(client, tmp_path):
    ids = _staged(client, [_run("Ruchi", "June", 2), _run("Ruchi", "June ratio 2", 2)])
    res = client.post("/api/report/group", json={"run_ids": ids, "label": "June"})
    assert res.status_code == 200
    assert res.json()["name"] == "Ruchi (June).xlsx"

    wb = openpyxl.load_workbook(tmp_path / "Ruchi (June).xlsx")
    assert len(wb["June"]["A"]) >= 4 + 2         # banner, header, 4 rows, footers
    wb.close()


def test_two_brands_never_merge_into_one_workbook(client):
    ids = _staged(client, [_run("Ruchi", "June", 1), _run("Cocola", "June", 1)])
    res = client.post("/api/report/group", json={"run_ids": ids})
    assert res.status_code == 400
    assert "different brands" in res.json()["detail"]


def test_batch_zips_one_workbook_per_group(client):
    ids = _staged(client, [_run("Ruchi", "June", 1), _run("Ruchi", "June ratio 2", 1),
                           _run("Cocola", "June", 1)])
    res = client.post("/api/report/batch", json={"groups": [
        {"run_ids": ids[:2], "label": "June"},
        {"run_ids": ids[2:]},
    ]})
    assert res.status_code == 200
    assert res.json()["workbooks"] == ["Ruchi (June).xlsx", "Cocola (June).xlsx"]


def test_a_flat_run_id_list_still_means_one_workbook_each(client):
    """The older callers and the CLI pass `run_ids`; each stays its own file."""
    ids = _staged(client, [_run("Ruchi", "June", 1), _run("Ruchi", "June ratio 2", 1)])
    res = client.post("/api/report/batch", json={"run_ids": ids})
    assert res.json()["workbooks"] == ["Ruchi (June).xlsx", "Ruchi (June ratio 2).xlsx"]


def test_a_merged_report_downloads_by_name(client):
    ids = _staged(client, [_run("Ruchi", "June", 1), _run("Ruchi", "June ratio 2", 1)])
    name = client.post("/api/report/group",
                       json={"run_ids": ids, "label": "June"}).json()["name"]
    res = client.get(f"/api/report/download/{name}")
    assert res.status_code == 200
    assert res.headers["content-type"].startswith("application/vnd.openxml")
