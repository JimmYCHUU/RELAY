"""Resolving Facebook cells the export could not account for on its own.

The sheet is hand-maintained: a caption is written, the links are pasted, and
later the caption is edited *on the post* without the sheet being updated. One
caption serves every FB slot in its row, so a stale one breaks the whole row.

It cannot be repaired offline — measured on a real June export, 32 of 40 such
rows score under 0.70 against every post on one of their pages, because the text
was rewritten rather than tweaked. So one post per row is read through the
browser and identified by its numeric post id, which needs no caption at all;
the row's other posts are then found from that anchor, because Somoy cross-posts
one story to several pages within minutes.

The browser is faked throughout — every rule here is decidable without one.
"""
from __future__ import annotations

from contextlib import contextmanager
from datetime import datetime

from relay.collectors.base import Pacer
from relay.ingest.insights import build_index
from relay.matching.normalize import strip_boilerplate
from relay.models import CellValue, ReportRow, RunResult
from relay.resolve.insights_fill import fill_from_insights, refresh_caption

MAIN = "https://www.facebook.com/somoynews.tv/posts/"
SPORT = "https://www.facebook.com/somoytvsports/posts/"
SHONGBAD = "https://www.facebook.com/somoysongbad360/posts/"

HEADER = ('"Post ID","Page ID","Page name",Title,"Publish time",Permalink,'
          '"Is share","Post type",Views,Reach,'
          '"Reactions, comments and shares",Reactions,Comments\n')

# What the sheet says, and what the post says after someone rewrote it.
SHEET_CAPTION = "এখন থেকে সব পাবলিক পরীক্ষায় সব বোর্ডে অভিন্ন প্রশ্ন থাকবে"
POST_CAPTION = "এখন থেকে সব বোর্ডের এসএসসি ও এইচএসসি পরীক্ষা অভিন্ন প্রশ্নে হবে"
JUNE3 = datetime(2026, 6, 3)


def _csv(tmp_path, rows: str, name="export.csv"):
    path = tmp_path / name
    path.write_text(HEADER + rows, encoding="utf-8-sig")
    return path


def _export(tmp_path, extra: str = "") -> object:
    """The rewritten story as all three pages carry it, minutes apart."""
    rows = (
        f'5001,100064517327464,"somoynews.tv","{POST_CAPTION}...\n\nবিস্তারিত কমেন্টে…\n\n#somoytv",'
        f'"06/03/2026 20:00",{MAIN}pfbid0EXPORTMAIN,0,Photos,182392,120000,90,70,9\n'
        f'5002,61550125032197,"Somoy Sports","{POST_CAPTION}...\n\nবিস্তারিত কমেন্টে…\n\n#sports",'
        f'"06/03/2026 20:12",{SPORT}pfbid0EXPORTSPORT,0,Photos,6542,5000,20,15,2\n'
        f'5003,900,"সময় সংবাদ","{POST_CAPTION}","06/03/2026 20:20",'
        f'{SHONGBAD}pfbid0EXPORTSHONGBAD,0,Photos,4311,3000,18,12,1\n'
    ) + extra
    return build_index([_csv(tmp_path, rows)])


def _row(no, caption, links, date=JUNE3):
    return ReportRow(
        no=no, date=date, caption=caption,
        links={"fb1": links.get("fb1"), "fb2": links.get("fb2"),
               "fb3": links.get("fb3"), "x": None, "ig": None},
        cells={s: CellValue.missing() for s in ("fb1", "fb2", "fb3", "x", "ig")})


def _run(rows, index=None):
    r = RunResult(brand="B", month="June", rows=rows)
    r.insights = index
    return r


def _stale_run(index, links=None):
    """One row the sheet describes with the caption the post no longer carries."""
    links = links or {"fb1": f"{MAIN}pfbid0SHEETSIDE", "fb2": f"{SPORT}pfbid0SHEETSIDE2"}
    return _run([_row(1, SHEET_CAPTION, links)], index)







def _patch(monkeypatch, *, post_id=None, by_url=None, caption=None, views=None):
    """Stand in for the browser. `by_url` maps a campaign link to the post id its
    page would report, since the real pass now visits each unresolved slot in
    turn and every post answers with its own id."""
    from relay.collectors import browser, mbs

    @contextmanager
    def fake_session(_profile, headed=False, recycle_every=None):
        yield type("S", (), {"page": lambda self: object()})()

    def visit(page, url, pacer):
        pacer.before_visit(url)
        pid = (by_url or {}).get(url, post_id)
        return (CellValue(views, "collected", 1.0, "post page views figure") if views
                else CellValue.missing("no views figure"), None, pid)

    monkeypatch.setattr(browser, "persistent_session", fake_session)
    monkeypatch.setattr(mbs, "extract_caption", lambda page: caption)
    monkeypatch.setattr(mbs, "collect_fb_post", visit)


def test_resolve_identifies_the_post_and_refreshes_the_caption(tmp_path, monkeypatch):
    from relay.collectors.runner import Progress, resolve_facebook

    index = _export(tmp_path)
    run = _stale_run(index)
    _patch(monkeypatch, post_id="5002")          # the Somoy Sports row, via FB2
    pacer, prog = Pacer(dry_run=False), Progress()
    monkeypatch.setattr(pacer, "before_visit", lambda url: None)

    filled = resolve_facebook(run, pacer=pacer, progress=prog)
    row = run.rows[0]
    assert row.cells["fb2"].value == 6542, "anchor filled by post id, no caption involved"
    assert row.caption == strip_boilerplate(POST_CAPTION), "the post's own wording"
    assert row.original_caption == SHEET_CAPTION, "the sheet's wording stays recoverable"
    assert filled == 2 and prog.state == "finished"


def test_the_rest_of_the_row_costs_no_second_visit(tmp_path, monkeypatch):
    """Somoy cross-posts one story to several pages minutes apart, so the anchor
    identifies the whole row. This is the entire reason one visit per row is enough."""
    from relay.collectors.runner import Progress, resolve_facebook

    index = _export(tmp_path)
    run = _stale_run(index, {"fb1": f"{MAIN}pfbid0A", "fb2": f"{SPORT}pfbid0B",
                             "fb3": f"{SHONGBAD}pfbid0C"})
    _patch(monkeypatch, post_id="5002")
    pacer = Pacer(dry_run=False)
    visits = []
    monkeypatch.setattr(pacer, "before_visit", lambda url: visits.append(url))

    filled = resolve_facebook(run, pacer=pacer, progress=Progress())
    row = run.rows[0]
    assert len(visits) == 1, "one post read, three cells filled"
    assert filled == 3
    assert row.cells["fb1"].value == 182392 and row.cells["fb3"].value == 4311
    assert "same story as the FB2 post" in row.cells["fb1"].note
    assert row.cells["fb1"].confidence <= 0.95, "the story matches; the URL never confirmed it"


def _siblings(run, index, anchor_post_id, anchor_slot="fb2"):
    """Run only the free step — the one that fills a row's other slots from an
    already-identified post, without navigating."""
    from relay.resolve.insights_fill import (_by_slug, _slot_pages,
                                             fill_row_from_anchor)
    row = run.rows[0]
    return fill_row_from_anchor(run, row, 0, anchor_slot,
                                index.by_post_id[anchor_post_id],
                                _by_slug(index), _slot_pages([run], index), index)


def test_a_page_running_the_story_twice_is_not_filled_for_free(tmp_path):
    """Two posts on one page inside the window carrying the same story: the link
    does not say which one it is, so the free step declines. The pass will still
    visit that slot afterwards and settle it by post id."""
    twin = (f'5004,100064517327464,"somoynews.tv","{POST_CAPTION}",'
            f'"06/03/2026 20:40",{MAIN}pfbid0EXPORTTWIN,0,Photos,99999,80000,50,40,4\n')
    index = _export(tmp_path, extra=twin)
    run = _stale_run(index, {"fb1": f"{MAIN}pfbid0A", "fb2": f"{SPORT}pfbid0B"})

    assert _siblings(run, index, "5002") == 0
    assert run.rows[0].cells["fb1"].value is None
    assert any("same story" in i.reason for i in run.issues)


def test_a_reversed_headline_is_not_the_same_story(tmp_path):
    """The repo's own regression pair: বাড়ানোর (raised) against কমানোর (cut).
    One word apart, a different post — and the exact shape that made every
    offline attempt at this problem produce wrong numbers instead of blanks."""
    raised = "দেশের বাজারে স্বর্ণের দাম আরেক দফা বাড়ানোর সিদ্ধান্ত নিয়েছে বাজুস"
    cut = "দেশের বাজারে আরেক দফা স্বর্ণের দাম কমানোর সিদ্ধান্ত নিয়েছে বাজুস"
    rows = (f'6001,61550125032197,"Somoy Sports","{raised}","06/03/2026 20:00",'
            f'{SPORT}pfbid0RAISED,0,Photos,6542,5000,20,15,2\n'
            f'6002,100064517327464,"somoynews.tv","{cut}","06/03/2026 20:10",'
            f'{MAIN}pfbid0CUT,0,Photos,637905,400000,300,250,30\n')
    index = build_index([_csv(tmp_path, rows)])
    run = _run([_row(1, "একটি সম্পূর্ণ ভিন্ন শিরোনাম যা কোথাও মেলে না",
                     {"fb1": f"{MAIN}pfbid0A", "fb2": f"{SPORT}pfbid0B"})], index)
    assert _siblings(run, index, "6001") == 0, "a cut is not a raise"
    assert run.rows[0].cells["fb1"].value is None


def test_an_unidentifiable_post_never_silently_rewrites_the_caption(tmp_path, monkeypatch):
    """og:description on a deleted or restricted post is Facebook's boilerplate
    or the page bio. Good enough to show a human, never to attach a number to."""
    from relay.collectors.runner import Progress, resolve_facebook

    index = _export(tmp_path)
    run = _stale_run(index)
    _patch(monkeypatch, post_id=None, caption=None)
    pacer = Pacer(dry_run=False)
    monkeypatch.setattr(pacer, "before_visit", lambda url: None)

    filled = resolve_facebook(run, pacer=pacer, progress=Progress())
    assert filled == 0
    assert run.rows[0].caption == SHEET_CAPTION, "untouched"
    assert run.rows[0].original_caption == ""
    assert any("could not be identified" in i.reason for i in run.issues)


def test_refresh_caption_is_idempotent():
    """A second repair must not overwrite the sheet's wording with the first
    repair's output — that is the only copy of it RELAY holds."""
    row = _row(1, SHEET_CAPTION, {"fb1": f"{MAIN}x"})
    assert refresh_caption(row, POST_CAPTION) == SHEET_CAPTION
    assert refresh_caption(row, "আরও একবার সম্পাদিত একটি ভিন্ন শিরোনাম") == POST_CAPTION
    assert row.original_caption == SHEET_CAPTION
    # an unchanged caption is not a repair
    assert refresh_caption(row, row.caption) == ""


# --- operational guards ---

def test_resolve_dry_run_launches_no_browser(tmp_path):
    from relay.collectors.runner import Progress, resolve_facebook

    index = _export(tmp_path)
    run = _stale_run(index)
    pacer, prog = Pacer(dry_run=True), Progress()
    assert resolve_facebook(run, pacer=pacer, progress=prog) == 0
    assert pacer.visits == prog.total == 1
    assert prog.state == "finished" and "dry-run" in prog.message
    assert run.rows[0].caption == SHEET_CAPTION


def test_resolve_without_an_export_does_nothing(tmp_path):
    from relay.collectors.runner import Progress, resolve_facebook

    run = _run([_row(1, SHEET_CAPTION, {"fb1": f"{MAIN}pfbid0A"})], None)
    prog = Progress()
    assert resolve_facebook(run, pacer=Pacer(dry_run=True), progress=prog) == 0
    assert prog.state == "finished" and "no insights export" in prog.message


def test_resolve_honours_the_stop_button(tmp_path, monkeypatch):
    from relay.collectors.runner import Progress, resolve_facebook

    index = _export(tmp_path)
    run = _stale_run(index)
    _patch(monkeypatch, post_id="5002")
    prog = Progress()
    prog.stop_requested = True
    pacer = Pacer(dry_run=False)
    monkeypatch.setattr(pacer, "before_visit", lambda url: None)

    assert resolve_facebook(run, pacer=pacer, progress=prog) == 0
    assert prog.state == "stopped" and prog.done == 0


def test_resolve_stops_when_the_pacing_budget_runs_out(tmp_path, monkeypatch):
    from relay.collectors.runner import Progress, resolve_facebook

    index = _export(tmp_path)
    rows = [_row(n, SHEET_CAPTION,
                 {"fb1": f"{MAIN}pfbid0A{n}", "fb2": f"{SPORT}pfbid0B{n}"})
            for n in (1, 2)]
    run = _run(rows, index)
    _patch(monkeypatch, post_id="5002")
    prog = Progress()

    assert resolve_facebook(run, pacer=Pacer(budget=1), progress=prog) == 2
    assert prog.state == "stopped"
    assert rows[0].cells["fb2"].value == 6542, "the first row still landed"
    assert rows[1].cells["fb2"].value is None


def test_the_workbook_never_explains_itself(tmp_path):
    """The delivered file goes to a sponsor who has no way to know how RELAY
    works and no reason to care. It carries the post's real caption and the
    figure — and not one word about where either came from."""
    import openpyxl

    from relay.report.generator import build_report

    row = _row(1, POST_CAPTION, {"fb1": f"{MAIN}pfbid0A"})
    row.original_caption = SHEET_CAPTION
    row.cells["fb1"] = CellValue(182392, "collected", 1.0,
                                 "insights export e.csv row 12 · post 5001 · Views")
    out = tmp_path / "report.xlsx"
    build_report(_run([row]), out)

    ws = openpyxl.load_workbook(out)[openpyxl.load_workbook(out).sheetnames[0]]
    assert ws["C3"].value == POST_CAPTION
    assert not [c for r in ws.iter_rows() for c in r if c.comment], "no annotations"


def test_every_unresolved_slot_is_visited_in_turn(tmp_path, monkeypatch):
    """The row that prompted this merge: its caption is the placeholder
    "PREDICTION", so nothing offline can identify any of its three posts. The
    old two-button flow visited one and left the other two blank. Each post
    answers with its own id, and each cell is settled from it."""
    from relay.collectors.runner import Progress, resolve_facebook

    # Three unrelated headlines, so the free same-story step can recover nothing
    # and every slot has to be visited on its own.
    rows = (f'7001,100064517327464,"somoynews.tv","প্রথম সম্পূর্ণ ভিন্ন একটি শিরোনাম",'
            f'"06/03/2026 20:00",{MAIN}pfbid0X1,0,Photos,182392,120000,90,70,9\n'
            f'7002,61550125032197,"Somoy Sports","দ্বিতীয় সম্পূর্ণ ভিন্ন একটি শিরোনাম",'
            f'"06/03/2026 20:05",{SPORT}pfbid0X2,0,Photos,6542,5000,20,15,2\n'
            f'7003,900,"সময় সংবাদ","তৃতীয় সম্পূর্ণ ভিন্ন একটি শিরোনাম",'
            f'"06/03/2026 20:10",{SHONGBAD}pfbid0X3,0,Photos,4311,3000,18,12,1\n')
    index = build_index([_csv(tmp_path, rows)])
    links = {"fb1": f"{MAIN}pfbid0A", "fb2": f"{SPORT}pfbid0B", "fb3": f"{SHONGBAD}pfbid0C"}
    run = _run([_row(1, "PREDICTION", links)], index)
    _patch(monkeypatch, by_url={links["fb1"]: "7001", links["fb2"]: "7002",
                                links["fb3"]: "7003"})
    pacer = Pacer(dry_run=False)
    visits = []
    monkeypatch.setattr(pacer, "before_visit", lambda url: visits.append(url))

    assert resolve_facebook(run, pacer=pacer, progress=Progress()) == 3
    row = run.rows[0]
    assert (row.cells["fb1"].value, row.cells["fb2"].value, row.cells["fb3"].value) \
        == (182392, 6542, 4311)
    assert len(visits) == 3, "one visit per slot, since none could be had for free"


def test_nothing_is_ever_estimated(tmp_path, monkeypatch):
    """A post the export cannot account for leaves a blank, whatever its
    reaction count. The multiplier that used to fill these was measured wrong
    for ~92% of posts."""
    from relay.collectors.runner import Progress, resolve_facebook

    index = _export(tmp_path)
    run = _run([_row(1, SHEET_CAPTION, {"fb1": f"{MAIN}pfbid0UNKNOWN"})], index)
    _patch(monkeypatch, post_id="not-in-the-export")
    pacer = Pacer(dry_run=False)
    monkeypatch.setattr(pacer, "before_visit", lambda url: None)

    assert resolve_facebook(run, pacer=pacer, progress=Progress()) == 0
    cell = run.rows[0].cells["fb1"]
    assert cell.value is None and cell.provenance == "missing"
    assert any("could not be identified" in i.reason for i in run.issues)


def test_one_visit_settles_a_whole_cross_posted_row(tmp_path, monkeypatch):
    """The counterpart: when the row's posts do carry the same story, one visit
    accounts for all three. This is what keeps a 300-row month affordable."""
    from relay.collectors.runner import Progress, resolve_facebook

    index = _export(tmp_path)
    links = {"fb1": f"{MAIN}pfbid0A", "fb2": f"{SPORT}pfbid0B", "fb3": f"{SHONGBAD}pfbid0C"}
    run = _run([_row(1, "PREDICTION", links)], index)
    _patch(monkeypatch, post_id="5002")
    pacer = Pacer(dry_run=False)
    visits = []
    monkeypatch.setattr(pacer, "before_visit", lambda url: visits.append(url))

    assert resolve_facebook(run, pacer=pacer, progress=Progress()) == 3
    assert len(visits) == 1, "the other two came free from the same story"
