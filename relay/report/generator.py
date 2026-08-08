"""Sponsor report writer — "modern theme" (SRS FR-20..FR-22, SDD 5.6).

Layout (columns, footer formulas, merges) still mirrors the hand-made
the client's own hand-made report as ground truth; the visual layer is the
modern style the analyst approved from `style-samples/style-3-modern-theme.xlsx`.

The report's chrome wears the sponsor's own colour, read off the campaign
sheet's brand-name and header rows (see `palette.py`): the banner, the header
row, the rule beneath it, and the Sum / Total / Average rows. The data rows
stay neutral — the zebra stripe, captions and links are what a reader is
actually reading, and tinting those buys nothing.

Only those two header rows are consulted. An earlier attempt took the sheet's
strongest colour from anywhere in the tab, which on the June sheets picked up a
highlighted cell's yellow (FFFF00) and produced an olive masthead. A sheet with
no brand colour at all — SMC Plus's June tracker fills those rows plain grey —
falls back to the approved teal, unchanged.

Nothing in the delivered file explains how RELAY works: no cell comments, no
provenance, no notes. The dashboard is where a figure is audited; this workbook
is what goes to the sponsor.
"""
from __future__ import annotations

from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import column_index_from_string, get_column_letter

from ..matching.permalink import is_share_link, post_token
from ..models import RunResult
from .palette import DEFAULT, derive

# style-3's exact palette, still the fallback for an unbranded sheet. The tint
# and band are the sample's own hand-picked values, not a formula's output —
# mixing the accent toward white lands a few points off on both.
ACCENT = "0C8F7C"
TINT = "E7F5F2"
BAND = "F5F8F7"
INK = "FF1F2937"
LINK_BLUE = "FF1155CC"

COL_WIDTHS = {
    "A": 16.7, "B": 21.7, "C": 31.3,
    "D": 30.0, "E": 15.0, "F": 15.0, "G": 17.0, "H": 15.0,   # FB link 1 + its figures
    "I": 30.0, "J": 15.0, "K": 15.0, "L": 17.0, "M": 15.0,   # FB link 2
    "N": 30.0, "O": 15.0, "P": 15.0, "Q": 17.0, "R": 15.0,   # FB link 3
    "S": 27.7, "T": 24.6,                            # X
    "U": 25.0, "V": 19.6,                            # Instagram
    "W": 20.0,                                       # Source tab (merged reports)
}
HEADERS = [
    "No", "Date", "Content's name",
    "Content's Link 1", "Views", "Reach", "Engagement", "Clicks",
    "Content's Link 2", "Views", "Reach", "Engagement", "Clicks",
    "Content's Link 3", "Views", "Reach", "Engagement", "Clicks",
    "X, Link 4", "Impressions",
    "Instagram", "Views",
]
# The link cell for each slot, in template column order.
LINK_COLS = {"fb1": 4, "fb2": 9, "fb3": 14, "x": 19, "ig": 21}
# The three Facebook column groups, left to right.
FB_COLUMNS = ("fb1", "fb2", "fb3")
# Figure columns per slot. Only Facebook carries all four: Meta's content export
# publishes Reach, "Reactions, comments and shares" and Total clicks per post,
# and neither X's public page nor the Instagram export offers an equivalent — so
# those slots keep their single column rather than showing three permanently
# empty ones.
METRIC_COLS = {
    "fb1": {"views": 5, "reach": 6, "engagement": 7, "clicks": 8},
    "fb2": {"views": 10, "reach": 11, "engagement": 12, "clicks": 13},
    "fb3": {"views": 15, "reach": 16, "engagement": 17, "clicks": 18},
    "x": {"views": 20},
    "ig": {"views": 22},
}
LAST_COL = 22
# Appended only when a report merges several tabs of one workbook — it says which
# tab a row came from, the one thing the merge would otherwise lose.
SOURCE_COL = 23
SOURCE_HEADER = "Source tab"


def _metric_cols(name: str) -> tuple[int, ...]:
    return tuple(m[name] for m in METRIC_COLS.values() if name in m)


VIEW_COLS = _metric_cols("views")
REACH_COLS = _metric_cols("reach")
ENGAGEMENT_COLS = _metric_cols("engagement")
CLICK_COLS = _metric_cols("clicks")
NUMERIC_COLS = tuple(sorted(VIEW_COLS + REACH_COLS + ENGAGEMENT_COLS + CLICK_COLS))

# The footer, in order: every figure's grand total, then every figure's average
# per content. Views lead both blocks — that is the headline the sponsor reads
# first — and the totals stay together rather than alternating with the
# averages, so a reader comparing two figures compares like with like.
#
# The averages divide by the row count, the same denominator "Average views per
# content" has always used. A row whose Facebook post was never made still
# counts, because the average answers "what did a piece of content earn", and
# leaving those rows out would quietly answer a different question for reach,
# engagement and clicks than for views.
FOOTER_METRICS = (
    ("Total views", "Average views per content", VIEW_COLS),
    ("Total reach (Facebook)", "Average reach per content (Facebook)", REACH_COLS),
    ("Total engagement (Facebook)",
     "Average engagement per content (Facebook)", ENGAGEMENT_COLS),
    ("Total clicks (Facebook)", "Average clicks per content (Facebook)", CLICK_COLS),
)
# Where a footer row's label stops and its merged value begins.
LABEL_SPAN = 3
VALUE_COL = LABEL_SPAN + 1

DATE_FMT = "d\\ mmm"
NUM_FMT = "#,##0"

# Excel rejects these in a sheet title and caps it at 31 characters; a brand or
# tab name is user-supplied text that has no reason to respect either.
_TITLE_BAD = str.maketrans({c: "-" for c in "[]:*?/\\"})


def _safe_title(name: str) -> str:
    return (name or "Report").translate(_TITLE_BAD).strip("'")[:31] or "Report"


def _names_a_post(url: str | None) -> bool:
    """Whether a link identifies a Facebook post at all.

    A share link does once resolved; a bare page profile, or an Instagram URL
    someone pasted into a Facebook column, never will — so it can only ever sit
    beside three empty figures.
    """
    return bool(url) and (is_share_link(url) or post_token(url) is not None)


def _fb_order(row) -> list[str]:
    """Which slot each Facebook column shows, left to right.

    The delivered columns are "Content's Link 1 / 2 / 3" — plain positions with
    no page attached to them — so a row whose main-page post was never made
    opened on an empty link and three empty figures, with the real post sitting
    in Link 2. A sponsor reads that as a hole in the report rather than as a
    post that was never published.

    Only the first column is closed up, and by a swap: the first slot that does
    name a post moves into Link 1, and whatever Link 1 held takes its place. So
    nothing is dropped — a link the sheet supplied is still delivered, just not
    at the head of the row — and the two later columns keep the sheet's own
    order. A gap between Link 2 and Link 3 is left alone; it is not what a
    reader's eye lands on first.

    The review screen deliberately keeps the sheet's order throughout, where
    FB 1 means the main page and which page a figure came from is the point.
    """
    order = list(FB_COLUMNS)
    if _names_a_post(row.link(order[0])):
        return order
    donor = next((i for i in range(1, len(order))
                  if _names_a_post(row.link(order[i]))), None)
    if donor is None:
        return order                          # nothing to promote
    order[0], order[donor] = order[donor], order[0]
    return order


def _col_sum(cols, row: int) -> str:
    """A total across non-adjacent columns — the four figures interleave per
    slot, so no single range can add up one metric."""
    return "=" + "+".join(f"{get_column_letter(c)}{row}" for c in cols)


# ── writer ────────────────────────────────────────────────────────────────────
def _write_cell(ws, row, col, value, font, fmt=None, fill=None, border=None,
                align=None):
    c = ws.cell(row=row, column=col, value=value)
    c.font = font
    c.alignment = align or Alignment(horizontal="center", vertical="center",
                                     wrap_text=True)
    if border is not None:
        c.border = border
    if fmt:
        c.number_format = fmt
    if fill is not None:
        c.fill = fill
    return c


def build_report(
    result: RunResult,
    out_path: str | Path,
    sheet_name: str | None = None,
) -> Path:
    # The sponsor's colour dresses the chrome; the data rows below stay neutral.
    pal = derive(result.accent) if result.accent else DEFAULT

    fill_acc = PatternFill("solid", fgColor="FF" + pal.accent)
    fill_tint = PatternFill("solid", fgColor="FF" + pal.tint)
    fill_band = PatternFill("solid", fgColor="FF" + BAND)   # data zebra: neutral
    lgray = Side(style="thin", color="FFD5DDDA")
    border = Border(left=lgray, right=lgray, top=lgray, bottom=lgray)
    header_border = Border(left=lgray, right=lgray, top=lgray,
                           bottom=Side(style="medium", color="FF" + pal.rule))
    no_border = Border()
    # Banner and footer ink flip to dark on a light brand colour — white 38pt on
    # TK Super Board's gold or SMC Plus's green is unreadable.
    f_banner = Font(name="Arial", size=38, bold=True, color="FF" + pal.on_accent)
    f_header = Font(name="Arial", size=12, bold=True, color=INK)
    f_data = Font(name="Arial", size=12, bold=True, color="FF111111")
    f_caption = Font(name="Arial", size=11, color=INK)
    f_link = Font(name="Arial", size=9, color=LINK_BLUE, underline="single")
    f_sum = Font(name="Arial", size=12, bold=True, color="FF" + pal.accent_text)
    f_foot = Font(name="Arial", size=16, bold=True, color="FF" + pal.on_accent)
    left = Alignment(horizontal="left", vertical="center", wrap_text=True)

    # A merged report is the only one whose rows come from more than one tab, so
    # it is the only one that needs a column saying which.
    merged = any(r.source_sheet for r in result.rows)
    last_col = SOURCE_COL if merged else LAST_COL
    last_letter = get_column_letter(last_col)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = _safe_title(sheet_name or result.month)
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A3"
    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToWidth = 1
    # 0 = as many pages tall as it takes. Squeezing the height to one page as
    # well shrank a merged 206-row report to an unreadable smear; the width is
    # the only dimension that has to fit, so the columns stay on one sheet and
    # the rows run on.
    ws.page_setup.fitToHeight = 0
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    # …with the banner and column headers repeated at the top of each page.
    ws.print_title_rows = "1:2"

    for col, width in COL_WIDTHS.items():
        if column_index_from_string(col) <= last_col:
            ws.column_dimensions[col].width = width

    # banner: BRAND — Month Year
    year = next((r.date.year for r in result.rows if r.date), None)
    title = f"{result.brand.upper()} — {result.month}" + (f" {year}" if year else "")
    ws.merge_cells(f"A1:{last_letter}1")
    for col in range(1, last_col + 1):
        _write_cell(ws, 1, col, title if col == 1 else None, f_banner,
                    fill=fill_acc, border=no_border)
    ws.row_dimensions[1].height = 52

    # header
    headers = HEADERS + ([SOURCE_HEADER] if merged else [])
    for col, text in enumerate(headers, start=1):
        _write_cell(ws, 2, col, text, f_header, fill=fill_tint,
                    border=header_border)
    ws.row_dimensions[2].height = 34

    # data
    first_data = 3
    for i, r in enumerate(result.rows):
        excel_row = first_data + i
        lines = max(1, -(-len(r.caption or "") // 38))
        ws.row_dimensions[excel_row].height = min(130, max(70, lines * 13 + 10))
        band = fill_band if i % 2 else None
        # Renumber sequentially: the source No is hand-filled and may repeat or
        # be blank; the delivered report must carry a clean, unique No.
        _write_cell(ws, excel_row, 1, i + 1, f_data, fill=band, border=border)
        _write_cell(ws, excel_row, 2, r.date, f_data,
                    fmt=DATE_FMT if r.date else None, fill=band, border=border)
        _write_cell(ws, excel_row, 3, r.caption, f_caption, fill=band,
                    border=border, align=left)
        # Each Facebook column shows whichever slot `_fb_order` puts there, and
        # that slot's four figures travel with its link. X and Instagram have a
        # column each and never move.
        shown = list(zip(FB_COLUMNS, _fb_order(r))) + [("x", "x"), ("ig", "ig")]
        for col_slot, slot in shown:
            lc = LINK_COLS[col_slot]
            link = r.link(slot)
            cell = r.cells[slot]
            link_cell = _write_cell(ws, excel_row, lc, link,
                                    f_link if link else f_data,
                                    fill=band, border=border)
            if link:
                link_cell.hyperlink = link
            for metric, vc in METRIC_COLS[col_slot].items():
                # Reach, engagement and clicks exist only for a cell the export
                # filled or a person typed; a collected figure leaves them blank
                # rather than implying a zero the post did not report.
                _write_cell(ws, excel_row, vc, getattr(cell, metric, None)
                            if metric != "views" else cell.value,
                            f_data, fmt=NUM_FMT, fill=band, border=border)
        if merged:
            _write_cell(ws, excel_row, SOURCE_COL, r.source_sheet, f_caption,
                        fill=band, border=border)

    n = len(result.rows)
    last_data = first_data + n - 1
    sum_row = last_data + 1
    first_total = sum_row + 1        # Total views — the headline, kept first
    first_avg = first_total + len(FOOTER_METRICS)

    # footer: Sum — one per figure column, so each column totals itself.
    ws.merge_cells(start_row=sum_row, start_column=1,
                   end_row=sum_row, end_column=LABEL_SPAN)
    for col in range(1, last_col + 1):
        _write_cell(ws, sum_row, col, "Sum" if col == 1 else None, f_sum,
                    fill=fill_tint, border=border)
    for col in NUMERIC_COLS:
        cl = get_column_letter(col)
        _write_cell(ws, sum_row, col,
                    f"=SUM({cl}{first_data}:{cl}{last_data})" if n else 0,
                    f_sum, fmt=NUM_FMT, fill=fill_tint, border=border)

    # footer: the four grand totals, then the same four as an average per content
    value_letter = get_column_letter(VALUE_COL)
    totals = [
        (first_total + i, label, _col_sum(cols, sum_row), NUM_FMT)
        for i, (label, _, cols) in enumerate(FOOTER_METRICS)
    ] + [
        # Each average divides its own total, one row apiece above. n == 0 would
        # make this =D../0 — a #DIV/0! in the sponsor's copy.
        (first_avg + i, label,
         f"={value_letter}{first_total + i}/{n}" if n else 0, "0")
        for i, (_, label, _cols) in enumerate(FOOTER_METRICS)
    ]
    for rr, label, formula, fmt in totals:
        ws.merge_cells(start_row=rr, start_column=1, end_row=rr, end_column=LABEL_SPAN)
        ws.merge_cells(start_row=rr, start_column=VALUE_COL, end_row=rr,
                       end_column=last_col)
        for col in range(1, last_col + 1):
            _write_cell(ws, rr, col, label if col == 1 else None, f_foot,
                        fill=fill_acc, border=no_border)
        _write_cell(ws, rr, VALUE_COL, formula, f_foot, fmt=fmt,
                    fill=fill_acc, border=no_border)
        ws.row_dimensions[rr].height = 28
    ws.row_dimensions[sum_row].height = 28

    out_path = Path(out_path)
    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)
    wb.close()
    return out_path
