"""Per-brand report palette, taken from the campaign sheet itself.

Every sponsor's campaign tracker fills its brand-name row and the header row
beneath it with that brand's colour — bKash magenta, TK Super Board gold,
Fresh LPG blue. The delivered report should wear the same colour, and in a
multi-brand cycle each workbook wears its own, so the palette travels on the
run rather than living as a module constant.

A sheet that was never branded fills those rows with plain grey (SMC Plus's
June tracker does). Grey is the absence of a brand colour, not a brand colour,
so anything that close to neutral falls back to the approved teal — which is
also why this reads only the two header rows. An earlier attempt took the
sheet's strongest colour anywhere and turned a highlighted cell into an olive
masthead.

Foreground is computed, never assumed: `F1C232` gold and `93C47D` green are
light enough that the banner's white text would be unreadable on them, so the
banner ink flips to dark by WCAG contrast.
"""
from __future__ import annotations

from dataclasses import dataclass
from pathlib import Path

import openpyxl

# The approved teal, used when a sheet carries no brand colour. Its tint and
# band are the style sample's own hand-picked values rather than this module's
# formula — mixing the accent toward white lands a couple of points off on
# both, and the sample is the approved look.
TEAL_ACCENT = "0C8F7C"
TEAL_TINT = "E7F5F2"
TEAL_BAND = "F5F8F7"

INK = "1F2937"          # dark foreground, matching the report's body ink
WHITE = "FFFFFF"

# Below this max-minus-min RGB spread a fill is grey, not a brand colour.
MIN_CHROMA = 24
# Near-white and near-black fills are page furniture, never an accent.
MIN_LUMA, MAX_LUMA = 12, 244


# ── colour maths ──────────────────────────────────────────────────────────────
def _rgb(hex6: str) -> tuple[int, int, int]:
    return int(hex6[0:2], 16), int(hex6[2:4], 16), int(hex6[4:6], 16)


def _hex(r: int, g: int, b: int) -> str:
    clamp = lambda v: max(0, min(255, int(round(v))))
    return f"{clamp(r):02X}{clamp(g):02X}{clamp(b):02X}"


def chroma(hex6: str) -> int:
    r, g, b = _rgb(hex6)
    return max(r, g, b) - min(r, g, b)


def _luma(hex6: str) -> float:
    """WCAG relative luminance."""
    out = []
    for c in _rgb(hex6):
        s = c / 255
        out.append(s / 12.92 if s <= 0.03928 else ((s + 0.055) / 1.055) ** 2.4)
    r, g, b = out
    return 0.2126 * r + 0.7152 * g + 0.0722 * b


def contrast(a: str, b: str) -> float:
    la, lb = _luma(a), _luma(b)
    hi, lo = max(la, lb), min(la, lb)
    return (hi + 0.05) / (lo + 0.05)


def mix_on_white(hex6: str, ratio: float) -> str:
    """`ratio` of the colour over white. 0.10 and 0.04 reproduce the tint and
    band of the approved style-6 samples exactly, for both its navy and its
    red."""
    r, g, b = _rgb(hex6)
    return _hex(255 - ratio * (255 - r),
                255 - ratio * (255 - g),
                255 - ratio * (255 - b))


def scale(hex6: str, factor: float) -> str:
    r, g, b = _rgb(hex6)
    return _hex(r * factor, g * factor, b * factor)


def readable_on(bg: str) -> str:
    """White or dark ink, whichever the eye can actually read on `bg`."""
    return WHITE if contrast(WHITE, bg) >= contrast(INK, bg) else INK


def _darken_until(fg: str, bg: str, target: float = 4.5) -> str:
    """Step a colour darker until it reads against `bg`. A gold accent as Sum-row
    text on its own pale tint is otherwise almost invisible."""
    out = fg
    for _ in range(24):
        if contrast(out, bg) >= target:
            return out
        nxt = scale(out, 0.88)
        if nxt == out:                     # bottomed out at black
            return out
        out = nxt
    return out


# ── palette ───────────────────────────────────────────────────────────────────
@dataclass(frozen=True)
class Palette:
    accent: str          # banner and footer fill
    tint: str            # header row and Sum row fill
    band: str            # zebra stripe on alternate data rows
    on_accent: str       # text drawn on `accent`
    accent_text: str     # the Sum row's coloured figures, drawn on `tint`
    source: str          # "campaign sheet" | "default"

    @property
    def rule(self) -> str:
        """The medium border under the header row."""
        return self.accent_text


def derive(accent: str, source: str = "campaign sheet") -> Palette:
    accent = accent.upper()
    tint = mix_on_white(accent, 0.10)
    band = mix_on_white(accent, 0.04)
    return Palette(
        accent=accent,
        tint=tint,
        band=band,
        on_accent=readable_on(accent),
        accent_text=_darken_until(scale(accent, 0.75), tint),
        source=source,
    )


DEFAULT = Palette(
    accent=TEAL_ACCENT,
    tint=TEAL_TINT,
    band=TEAL_BAND,
    on_accent=readable_on(TEAL_ACCENT),
    accent_text=TEAL_ACCENT,
    source="default",
)


# ── extraction ────────────────────────────────────────────────────────────────
def _theme_rgbs(wb) -> list[str]:
    """The workbook theme's colour scheme, in the index order Excel's
    `theme="n"` attribute uses."""
    import re
    raw = getattr(wb, "loaded_theme", None)
    if not raw:
        return []
    xml = raw.decode("utf-8", "ignore") if isinstance(raw, bytes) else str(raw)
    scheme = re.search(r"<a:clrScheme.*?</a:clrScheme>", xml, re.S)
    if not scheme:
        return []
    found: dict[str, str] = {}
    for m in re.finditer(
        r"<a:(dk1|lt1|dk2|lt2|accent[1-6]|hlink|folHlink)>(.*?)</a:\1>",
        scheme.group(0), re.S,
    ):
        name, body = m.group(1), m.group(2)
        val = re.search(r'(?:srgbClr|sysClr[^>]*?lastClr)="?val="?([0-9A-Fa-f]{6})', body)
        if not val:
            val = re.search(r'val="([0-9A-Fa-f]{6})"', body)
        if val:
            found[name] = val.group(1).upper()
    order = ["lt1", "dk1", "lt2", "dk2", "accent1", "accent2", "accent3",
             "accent4", "accent5", "accent6", "hlink", "folHlink"]
    return [found.get(k, "") for k in order]


def _apply_tint(hex6: str, tint: float) -> str:
    """Excel's tint: positive lightens toward white, negative darkens."""
    if not tint:
        return hex6
    r, g, b = _rgb(hex6)
    if tint > 0:
        f = lambda c: c + (255 - c) * tint
    else:
        f = lambda c: c * (1 + tint)
    return _hex(f(r), f(g), f(b))


def _fill_hex(cell, theme: list[str]) -> str | None:
    fill = cell.fill
    if fill is None or fill.fill_type != "solid":
        return None
    c = fill.fgColor
    if c is None:
        return None
    if getattr(c, "type", None) == "rgb" and isinstance(c.rgb, str):
        # openpyxl hands back ARGB; some sheets write alpha 00 for an opaque fill
        if len(c.rgb) == 8:
            return c.rgb[2:].upper()
        if len(c.rgb) == 6:
            return c.rgb.upper()
        return None
    if getattr(c, "type", None) == "theme":
        idx = c.theme
        if isinstance(idx, int) and 0 <= idx < len(theme) and theme[idx]:
            return _apply_tint(theme[idx], c.tint or 0)
    return None


def is_brand_colour(hex6: str) -> bool:
    if chroma(hex6) < MIN_CHROMA:
        return False
    r, g, b = _rgb(hex6)
    avg = (r + g + b) / 3
    return MIN_LUMA <= avg <= MAX_LUMA


def accent_from_campaign(path: str | Path, sheet: str) -> str | None:
    """The brand's colour, read from the campaign sheet's brand-name and header
    rows. Returns None when the sheet was never branded."""
    try:
        wb = openpyxl.load_workbook(path)      # styles need data_only=False
    except Exception:
        return None
    try:
        if sheet not in wb.sheetnames:
            return None
        ws = wb[sheet]
        theme = _theme_rgbs(wb)

        header = None
        for row in ws.iter_rows(min_row=1, max_row=10):
            a = row[0].value
            joined = " ".join(str(c.value) for c in row[:8] if c.value)
            if isinstance(a, str) and a.strip().lower() == "no" and "content" in joined.lower():
                header = row[0].row
                break
        last = header or 2

        # Count every branded fill across the banner and header rows; the
        # header row is filled right across, so the brand colour dominates.
        tally: dict[str, int] = {}
        for r in range(1, last + 1):
            for cell in ws[r][:13]:
                hx = _fill_hex(cell, theme)
                if hx and is_brand_colour(hx):
                    tally[hx] = tally.get(hx, 0) + 1
        if not tally:
            return None
        return max(tally, key=lambda k: (tally[k], -list(tally).index(k)))
    except Exception:
        return None
    finally:
        wb.close()


def palette_for(path: str | Path, sheet: str) -> Palette:
    accent = accent_from_campaign(path, sheet)
    return derive(accent) if accent else DEFAULT
