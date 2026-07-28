"""Meta Business Suite content-export parser.

The export is the only source of *exact* Facebook figures RELAY has. Verified
against a real half-month export: every row carried both Views and Reach,
including the `Is share = 1` rows — which is why shared posts no longer need
estimating.

Two format details the sample forced:
  * the CSV is UTF-8 **with BOM**, so `utf-8-sig` is required or the first
    header becomes "﻿Post ID";
  * `Title` holds the post caption verbatim, embedded newlines and all, so the
    file must go through the `csv` module rather than naive line splitting —
    records routinely span several physical lines each.

Header names are matched tolerantly because Meta renames and localizes them;
the alias table lives in `config.INSIGHTS_HEADERS` next to `SELECTORS`, which
exists for exactly this kind of fragile external identifier.
"""
from __future__ import annotations

import csv
import logging
import re
from dataclasses import dataclass, field
from datetime import datetime
from pathlib import Path

from .. import config
from ..matching.permalink import (ig_shortcode, normalize_fb_url, page_slug,
                                  post_token)
from ..models import InsightsRow, RowIssue

log = logging.getLogger("relay.ingest")

_WS = re.compile(r"\s+")
_TRUTHY = {"1", "true", "yes", "y"}
# Dashboard uploads are stored under a random per-upload prefix
# ("48bcbfb9_combine.csv"), which is not the filename anyone was handed.
_UPLOAD_PREFIX = re.compile(r"^[0-9a-f]{8}_(.+)$")
# Meta writes "07/15/2026 00:14"; other locales/exports vary, so try a short list.
_TIME_FORMATS = ("%m/%d/%Y %H:%M", "%m/%d/%Y", "%Y-%m-%d %H:%M:%S", "%Y-%m-%d %H:%M",
                 "%Y-%m-%d", "%d/%m/%Y %H:%M", "%d/%m/%Y")


def display_name(name: str) -> str:
    """An export's filename as the client knows it, for anything a human reads.

    A figure is only auditable if the file it cites is one the reader can find;
    the upload prefix makes the note point at a name that exists nowhere but
    RELAY's own uploads directory.
    """
    hit = _UPLOAD_PREFIX.match(name or "")
    return hit.group(1) if hit else (name or "")


def _norm_header(name: object) -> str:
    return _WS.sub(" ", str(name or "").replace("﻿", "").strip()).lower()


def _to_int(v: object) -> int | None:
    if v is None or isinstance(v, bool):
        return None
    if isinstance(v, int):
        return v
    if isinstance(v, float):
        return int(v)
    s = str(v).strip().replace(",", "")
    if not s or s.upper() in ("N/A", "NA", "-", "—"):
        return None
    return int(float(s)) if re.fullmatch(r"-?\d+(\.\d+)?", s) else None


def _to_dt(v: object) -> datetime | None:
    if isinstance(v, datetime):
        return v
    s = str(v or "").strip()
    if not s:
        return None
    for fmt in _TIME_FORMATS:
        try:
            return datetime.strptime(s, fmt)
        except ValueError:
            continue
    return None


def resolve_headers(headers: list[str]) -> dict[str, int]:
    """Map logical field -> column index, exact aliases first.

    Exact-before-substring matters: "Reactions, comments and shares" sits next
    to "Reactions" in the real export, and a naive substring scan for
    "reactions" grabs the wrong one. A column already claimed exactly is also
    never re-used by a later substring pass.
    """
    norm = [_norm_header(h) for h in headers]
    out: dict[str, int] = {}
    claimed: set[int] = set()

    for fieldname, aliases in config.INSIGHTS_HEADERS.items():
        for alias in aliases:
            if alias in norm:
                idx = norm.index(alias)
                out[fieldname] = idx
                claimed.add(idx)
                break

    for fieldname, aliases in config.INSIGHTS_HEADERS.items():
        if fieldname in out:
            continue
        for alias in aliases:
            hit = next((i for i, h in enumerate(norm)
                        if i not in claimed and alias in h), None)
            if hit is not None:
                out[fieldname] = hit
                claimed.add(hit)
                break
    return out


@dataclass
class InsightsExport:
    """One parsed export file."""
    path: str
    rows: list[InsightsRow] = field(default_factory=list)
    issues: list[RowIssue] = field(default_factory=list)


@dataclass
class InsightsIndex:
    """Lookup across every export the user supplied, keyed by permalink.

    `by_url` is the primary key; `by_token` is a fallback for the case where
    the same post is addressed through a different page slug (vanity name in
    the campaign sheet, numeric page id in the export, or vice versa).
    """
    by_url: dict[str, InsightsRow] = field(default_factory=dict)
    by_token: dict[str, InsightsRow] = field(default_factory=dict)
    # Meta's numeric post id — the only identifier shared between a live post
    # page and this export. `pfbid` blobs differ between the two, and a
    # mainpage post's headline is often rewritten so captions won't match it
    # either, so for those rows this is the only key that works.
    by_post_id: dict[str, InsightsRow] = field(default_factory=dict)
    # Numeric page id -> the vanity slug that page's permalinks use. A campaign
    # sheet's `/photo/?fbid=…&set=pb.<page-id>` links name the page by id only,
    # and the export is the one place both spellings appear side by side.
    page_slugs: dict[str, str] = field(default_factory=dict)
    # Instagram shortcode -> row. The exact opposite of Facebook's situation: the
    # code in a copied Instagram link is the code in the export, so this join
    # needs no browser and no caption.
    by_shortcode: dict[str, InsightsRow] = field(default_factory=dict)
    rows: list[InsightsRow] = field(default_factory=list)
    files: list[str] = field(default_factory=list)
    issues: list[RowIssue] = field(default_factory=list)

    def add(self, export: InsightsExport) -> None:
        self.files.append(export.path)
        self.issues.extend(export.issues)
        for row in export.rows:
            self.rows.append(row)
            url = normalize_fb_url(row.permalink)
            if url:
                self.by_url.setdefault(url, row)
            tok = post_token(row.permalink)
            if tok:
                self.by_token.setdefault(tok, row)
            if row.post_id:
                self.by_post_id.setdefault(row.post_id, row)
            slug = page_slug(row.permalink)
            if row.page_id and slug:
                self.page_slugs.setdefault(row.page_id, slug)
            code = ig_shortcode(row.permalink)
            if code:
                self.by_shortcode.setdefault(code, row)

    def slug_for_page_id(self, page_id: str | None) -> str | None:
        return self.page_slugs.get(str(page_id)) if page_id else None

    def lookup_ig(self, url: str | None) -> InsightsRow | None:
        """Join an Instagram campaign link to its export row by shortcode."""
        code = ig_shortcode(url)
        return self.by_shortcode.get(code) if code else None

    def lookup(self, url: str | None) -> InsightsRow | None:
        key = normalize_fb_url(url)
        if key and key in self.by_url:
            return self.by_url[key]
        tok = post_token(url)
        if tok:
            return self.by_token.get(tok)
        return None

    def lookup_post_id(self, post_id: str | None) -> InsightsRow | None:
        """Join on Meta's numeric post id, read off the live post page by
        `collectors.mbs.collect_fb_post`. Exact — no similarity involved."""
        return self.by_post_id.get(str(post_id)) if post_id else None

    def __len__(self) -> int:
        return len(self.rows)


def _build_row(values: list[object], cols: dict[str, int], fname: str,
               rownum: int) -> InsightsRow | None:
    def cell(name: str) -> object:
        idx = cols.get(name)
        return values[idx] if idx is not None and idx < len(values) else None

    permalink = str(cell("permalink") or "").strip()
    if not permalink:
        return None
    return InsightsRow(
        post_id=str(cell("post_id") or "").strip(),
        permalink=permalink,
        page_id=str(cell("page_id") or "").strip(),
        page_name=str(cell("page_name") or "").strip(),
        title=str(cell("title") or "").strip(),
        published=_to_dt(cell("published")),
        views=_to_int(cell("views")),
        reach=_to_int(cell("reach")),
        reactions=_to_int(cell("reactions")),
        post_type=str(cell("post_type") or "").strip(),
        is_share=str(cell("is_share") or "").strip().lower() in _TRUTHY,
        source_file=fname,
        source_row=rownum,
    )


def _read_csv(path: Path) -> tuple[list[str], list[list[object]]]:
    # utf-8-sig strips the BOM Meta writes; newline="" lets csv handle the
    # newlines embedded in Bengali captions.
    with path.open(encoding="utf-8-sig", newline="") as fh:
        reader = csv.reader(fh)
        rows = [r for r in reader]
    if not rows:
        return [], []
    return rows[0], [list(r) for r in rows[1:]]


def _read_xlsx(path: Path) -> tuple[list[str], list[list[object]]]:
    import openpyxl

    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    try:
        ws = wb[wb.sheetnames[0]]
        it = ws.iter_rows(values_only=True)
        header = next(it, None)
        if not header:
            return [], []
        return [str(h or "") for h in header], [list(r) for r in it]
    finally:
        wb.close()


def parse_insights(path: str | Path) -> InsightsExport:
    """Parse one Business Suite content export (.csv or .xlsx)."""
    p = Path(path)
    fname = p.name
    headers, raw_rows = _read_csv(p) if p.suffix.lower() == ".csv" else _read_xlsx(p)
    if not headers:
        raise ValueError(f"{fname}: file is empty")

    cols = resolve_headers(headers)
    missing = [f for f in ("permalink", "reach", "views") if f not in cols]
    if "permalink" in missing:
        raise ValueError(
            f"{fname}: no Permalink column found — headers were {headers[:12]}")

    export = InsightsExport(path=str(p))
    if missing:
        export.issues.append(RowIssue(
            fname, 1, f"export has no {'/'.join(missing)} column; "
                      "those rows cannot fill cells"))

    for rownum, values in enumerate(raw_rows, start=2):
        if not any(str(v or "").strip() for v in values):
            continue
        row = _build_row(values, cols, fname, rownum)
        if row is None:
            export.issues.append(RowIssue(fname, rownum, "row has no permalink — skipped"))
            continue
        export.rows.append(row)

    log.info("parsed %s: %d posts", fname, len(export.rows))
    return export


def build_index(paths: list[str | Path] | None) -> InsightsIndex:
    """Parse every supplied export into one lookup.

    A file that fails to parse is recorded as an issue rather than aborting the
    run — one bad export must never cost the user the whole report.
    """
    index = InsightsIndex()
    for path in paths or []:
        try:
            index.add(parse_insights(path))
        except Exception as exc:
            log.warning("insights export %s failed to parse: %s", path, exc)
            index.issues.append(RowIssue(Path(path).name, 0, f"export unreadable: {exc}"))
    return index
