"""Supervisor matched-file parser (SRS FR-2/FR-3).

Observed layouts, all with `Title` in column A:
  mainpage : Title | Views | Views_Match_1..N
  subpage  : Title | [Views |] Views_Match_1..N
  instagram: Title | Views_Match_1 | Views
Multi-brand files carry bare brand-token separator rows ("bkash", "White Plus").
"""
from __future__ import annotations

import re
from pathlib import Path

import openpyxl

from ..matching.normalize import is_brand_separator, normalize_caption
from ..models import MatchedFile, MatchedRow, RowIssue

_VM = re.compile(r"^views[_ ]?match[_ ]?(\d+)$", re.I)


def _to_int(v) -> int | None:
    if v is None or isinstance(v, bool):
        return None
    if isinstance(v, (int, float)):
        return int(v)
    s = str(v).strip().replace(",", "")
    return int(float(s)) if re.fullmatch(r"-?\d+(\.\d+)?", s) else None


def parse_matched(path: str | Path) -> MatchedFile:
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    try:
        ws = wb[wb.sheetnames[0]]
        fname = Path(path).name
        it = ws.iter_rows(values_only=True)
        header = next(it, None)
        if not header or str(header[0]).strip().lower() != "title":
            raise ValueError(f"{fname}: expected 'Title' header in A1, got {header!r}")

        vm_cols: list[tuple[int, int]] = []          # (match_index, column_index)
        for idx, name in enumerate(header[1:], start=1):
            if name is None:
                continue
            m = _VM.match(str(name).strip())
            if m:
                vm_cols.append((int(m.group(1)), idx))
        vm_cols.sort()
        if not vm_cols:
            raise ValueError(f"{fname}: no Views_Match_* columns found")

        rows: list[MatchedRow] = []
        sections: dict[str, list[MatchedRow]] = {}
        issues: list[RowIssue] = []
        current: list[MatchedRow] | None = None

        for rownum, raw in enumerate(it, start=2):
            title = raw[0] if raw else None
            if title is None or not str(title).strip():
                continue
            title = str(title)
            values = [_to_int(raw[c]) if c < len(raw) else None for _, c in vm_cols]
            while values and values[-1] is None:
                values.pop()

            if is_brand_separator(title, has_values=bool(values)):
                key = normalize_caption(title).lower()
                current = sections.setdefault(key, [])
                continue

            row = MatchedRow(title=title, values=values, source_row=rownum)
            rows.append(row)
            if current is not None:
                current.append(row)

        return MatchedFile(path=str(path), rows=rows, sections=sections, issues=issues)
    finally:
        wb.close()
