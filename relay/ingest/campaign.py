"""Campaign workbook parser (SRS FR-1).

Layout (verified across one sponsor's Feb/Mar/Apr/May/FifaWC/Election tabs):
row 1: free-form note cells; row 2: header
`No | Date | Content's name | Content's Link [1] | Content's Link 2 |
 Content's Link 3 | X, Link 4 | Instagram`; data until trailing empty region.
"""
from __future__ import annotations

from datetime import datetime
from pathlib import Path

import openpyxl

from ..models import CampaignRow, RowIssue

_EMPTY_RUN_STOP = 15  # consecutive empty rows = end of data (sheets have ~900 styled blanks)

# What sponsors call the row-number column. Every one of these sits in front of
# the same `Date | Content's name | Content's Link …` run, so the column's name
# is the least reliable thing about the header — Grameenphone's own workbook
# says "No" on its April tab and " SL" on May through August, and requiring the
# literal "No" left five months of that sponsor unreadable.
# Compared after dots and spaces are stripped, so "S. No" and "sl." land here too.
_NO_ALIASES = {"no", "sl", "slno", "serial", "serialno", "sr", "srno",
               "sno", "sn", "#", "id", "index"}


def list_sheets(path: str | Path) -> list[str]:
    wb = openpyxl.load_workbook(path, read_only=True)
    names = wb.sheetnames
    wb.close()
    return names


def _find_header(ws) -> tuple[int, dict] | None:
    """The header row and what each of its columns holds.

    The row is recognised by what follows the first column, not by the first
    column itself: "Content's Link" beside a "Date" is a campaign header and
    nothing else in these workbooks looks like one. The first cell then only has
    to be plausible as a row-number heading — one of `_NO_ALIASES`, or blank,
    which several tabs leave it.
    """
    for row in ws.iter_rows(min_row=1, max_row=10, max_col=_HEADER_SCAN_COLS):
        cells = [str(c.value).strip() if c.value is not None else "" for c in row]
        joined = " ".join(c for c in cells if c).lower()
        if "content" not in joined or "date" not in joined:
            continue
        first = cells[0].lower().replace(".", "").replace(" ", "")
        # A named first column is evidence in itself; a blank one leans entirely
        # on the Date + Content pair above, which is why that pair is required.
        if not first or first in _NO_ALIASES:
            return row[0].row, (_columns(cells) or dict(_FIXED))
    return None


"""Column roles, keyed to the header's own words rather than to a position.

`Content's Link 4 X` and `X, Link 4` are the same column under two spellings, and
both contain "link", so X and Instagram are claimed before the generic link
sweep sees them.
"""
_HEADER_SCAN_COLS = 14      # wide enough for the trailing Total Count / notes


def _is_x_header(h: str) -> bool:
    return (h == "x" or h.startswith("x,") or h.startswith("x ")
            or h.endswith(" x") or ", x" in h or "twitter" in h)


def _columns(cells: list[str]) -> dict | None:
    """Which column holds what, or None when the header does not say.

    Positions are not fixed across sponsors. Tecno POVA's June tab leaves one
    blank column between Link 1 and Link 2, and reading by position shifted
    every column after it along by one: the sheet's Link 3 was taken for the X
    post, its X post for the Instagram one, and the real Instagram column — the
    last of the row — was never read at all. So a Facebook URL was delivered in
    the report's X column, and 30 Instagram cells came back empty against an
    export that had every one of them.
    """
    out: dict = {"date": None, "caption": None, "x": None, "ig": None, "fb": []}
    for i, raw in enumerate(cells):
        h = raw.lower().strip()
        if not h:
            continue
        if out["date"] is None and h.startswith("date"):
            out["date"] = i
        elif out["caption"] is None and "content" in h and "name" in h:
            out["caption"] = i
        elif out["ig"] is None and "instagram" in h:
            out["ig"] = i
        elif out["x"] is None and _is_x_header(h):
            out["x"] = i
        elif "link" in h and len(out["fb"]) < 3:
            out["fb"].append(i)
    # Anything less is a header this function has misread, and guessing from
    # half a mapping is worse than the fixed positions that got us here.
    if out["caption"] is None or len(out["fb"]) < 2:
        return None
    return out


# The layout every sponsor but one actually uses, and the fallback for a header
# `_columns` cannot make sense of.
_FIXED = {"date": 1, "caption": 2, "fb": [3, 4, 5], "x": 6, "ig": 7}


def _cell_str(v) -> str | None:
    if v is None:
        return None
    s = str(v).strip()
    return s or None


def _link_str(v) -> str | None:
    """A link column counts as a link only when it holds an actual URL —
    the sheet's Sum/Total/Average footer rows park numbers in these columns."""
    s = _cell_str(v)
    if s and s.lower().startswith(("http://", "https://", "www.")):
        return s
    return None


def parse_campaign(path: str | Path, sheet: str) -> tuple[list[CampaignRow], list[RowIssue]]:
    wb = openpyxl.load_workbook(path, data_only=True)
    try:
        if sheet not in wb.sheetnames:
            raise ValueError(
                f"Sheet '{sheet}' not found in {Path(path).name}; available: {wb.sheetnames}"
            )
        ws = wb[sheet]
        found = _find_header(ws)
        if found is None:
            raise ValueError(f"Header row not found in sheet '{sheet}' of {Path(path).name}")
        header, cols = found
        # Only as far as the columns the header actually named — a tab can carry
        # a "Total Count" and a block of notes past its last link column.
        last_col = max([c for c in (cols["date"], cols["caption"], cols["x"],
                                    cols["ig"], *cols["fb"]) if c is not None]) + 1

        def at(row, key):
            """One mapped cell of a data row, or None when the tab has no such
            column — Grameenphone's tabs carry no Instagram link at all."""
            i = cols[key]
            return row[i].value if i is not None and i < len(row) else None

        rows: list[CampaignRow] = []
        issues: list[RowIssue] = []
        fname = Path(path).name
        empty_run = 0

        for row in ws.iter_rows(min_row=header + 1, max_col=last_col):
            r = row[0].row
            no, date = row[0].value, at(row, "date")
            caption = _cell_str(at(row, "caption"))
            links = [_link_str(row[i].value) if i < len(row) else None
                     for i in cols["fb"]]
            links += [None] * (3 - len(links))       # a tab with only two link columns
            x_link, ig_link = _link_str(at(row, "x")), _link_str(at(row, "ig"))

            if not caption and not any(links) and not x_link and not ig_link:
                empty_run += 1
                if empty_run >= _EMPTY_RUN_STOP:
                    break
                continue
            empty_run = 0

            # Some sheets keep an internal category-count row right under the
            # header ("Category A | 35 | Category B | 35 | …"): text/numbers parked
            # in the link columns, no Date, no URL anywhere. Company-side
            # bookkeeping, not campaign data — drop it from the run.
            #
            # Its No cell is blank on some sheets and holds the first category's
            # own name on others ("GP Bundle", "1 Poisha/Sec>>"), so the test is
            # that No is not a number rather than that it is empty — otherwise
            # Grameenphone's every tab opened with a content row captioned "46".
            link_junk = any(_cell_str(row[i].value) for i in range(3, len(row)))
            numbered = isinstance(no, (int, float)) and not isinstance(no, bool)
            if (not numbered and not isinstance(date, datetime)
                    and link_junk and not any(links) and not x_link and not ig_link):
                issues.append(RowIssue(
                    fname, r, "category-count row ignored (text in link columns, no URL/No/Date)",
                    where=f"sheet row {r}"))
                continue

            if not caption:
                # a link is trackable on its own: the row stays, caption-matching
                # is skipped, and collectors/manual entry fill the views —
                # collectors also recover the caption from the post itself
                issues.append(RowIssue(
                    fname, r, "caption empty — row kept; views fill from the post links",
                    where=f"sheet row {r}"))
                caption = ""

            if isinstance(no, float):
                no = int(no)
            if not isinstance(no, int):
                if no is not None:
                    issues.append(RowIssue(fname, r, f"non-numeric No '{no}' ignored",
                                           where=f"sheet row {r}"))
                no = None
            if not isinstance(date, datetime):
                if date is not None:
                    issues.append(RowIssue(fname, r, f"unreadable Date '{date}'",
                                           where=f"sheet row {r}"))
                date = None

            rows.append(CampaignRow(no, date, caption, links, x_link, ig_link, source_row=r))

        for i, cr in enumerate(rows):
            if cr.date is None:
                # posts are chronological, so a blank Date inherits a neighbour's:
                # previous row first, else the next row that has one
                fill = next((rows[j].date for j in range(i - 1, -1, -1) if rows[j].date), None) \
                    or next((rows[j].date for j in range(i + 1, len(rows)) if rows[j].date), None)
                if fill is not None:
                    cr.date = fill
                    # `{fill.day}` rather than strftime's `%-d`: the no-pad
                    # modifier is a glibc extension the Windows C runtime
                    # rejects outright with "Invalid format string", which the
                    # dashboard then reported as a bad campaign sheet.
                    issues.append(RowIssue(
                        fname, cr.source_row,
                        f"empty Date filled from adjacent row ({fill.day} {fill:%b})",
                        where=f"sheet row {cr.source_row}"))
                else:
                    issues.append(RowIssue(fname, cr.source_row, "missing Date",
                                           where=f"sheet row {cr.source_row}"))
            if cr.no is None:
                cr.no = (rows[i - 1].no + 1) if i and rows[i - 1].no is not None else i + 1
        return rows, issues
    finally:
        wb.close()
